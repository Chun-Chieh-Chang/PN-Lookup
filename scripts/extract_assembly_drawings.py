import os, sys, fitz, json, re, io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

print("=== 啟動組件圖面 PDF 結構化數據全量提取作業 (Assembly Drawing Pipeline) ===")

BASE_DIR = r"D:\Self-developed_Apps\PN-Lookup\rawdata\Drawings\組件"
JSON_OUTPUT = r"D:\Self-developed_Apps\PN-Lookup\data\assembly_drawings_extract.json"
XLSX_OUTPUT = r"D:\Self-developed_Apps\PN-Lookup\data\assembly_drawings_extract.xlsx"

# 1. 載入主資料庫作為交叉比對與補強基準
with open('data/pn-lookup-master.json', 'r', encoding='utf-8') as f:
    master = json.load(f)
master_parts = {p['partNo'].strip().upper(): p for p in master.get('parts', []) if 'partNo' in p}

# 2. 載入 icu-parts
icu_map = {}
if os.path.exists('data/icu-parts.json'):
    with open('data/icu-parts.json', 'r', encoding='utf-8') as f:
        for ip in json.load(f):
            pn = ip.get('partNo', '').strip().upper()
            if pn: icu_map[pn] = ip

# 3. 載入 drawings_extract_v7 作為既有成果索引
v7_drawings_map = {}
if os.path.exists('data/drawings_extract_v7.json'):
    with open('data/drawings_extract_v7.json', 'r', encoding='utf-8') as f:
        for it in json.load(f).get('items', []):
            fn = it.get('fileName', '')
            if fn: v7_drawings_map[fn.lower()] = it

# 色彩對照關鍵字
COLOR_KEYWORDS = [
    (r'\bBLUE\s+TINT\b|透明藍', 'Blue Tint (透明藍)'),
    (r'\bOPAQUE\s+BLUE\b|不透明藍', 'Opaque Blue (不透明藍)'),
    (r'\bLIGHT\s+BLUE\b|淺藍', 'Light Blue (淺藍)'),
    (r'\bDARK\s+BLUE\b|深藍', 'Dark Blue (深藍)'),
    (r'\bOPAQUE\s+WHITE\b|不透明白', 'Opaque White (不透明白)'),
    (r'\bOFF-WHITE\b|米白', 'Off-White (米白)'),
    (r'\bDARK\s+GREEN\b|深綠', 'Dark Green (深綠)'),
    (r'\bLIGHT\s+GREEN\b|淺綠', 'Light Green (淺綠)'),
    (r'\bWHITE\b|\bWHITE\s+PC\b|白', 'White (白)'),
    (r'\bBLACK\b|黑', 'Black (黑)'),
    (r'\bBLUE\b|\bBLUE\s+PC\b|藍', 'Blue (藍)'),
    (r'\bRED\b|紅', 'Red (紅)'),
    (r'\bGREEN\b|綠', 'Green (綠)'),
    (r'\bCLEAR\b|透明', 'Clear (透明)'),
    (r'\bNATURAL\b|本色|原色', 'Natural (本色/原色)'),
    (r'\bTRANSPARENT\b', 'Transparent (透明)'),
    (r'\bYELLOW\b|黃', 'Yellow (黃)'),
    (r'\bW[-_]?\d{3}\b', 'White (白色母)'),
    (r'\bK[-_]?\d{3}\b', 'Blue (藍色母)'),
]

PN_TOKEN_RE = re.compile(r'\b(?:[A-Z]{1,4}\d{1,4}(?:-\d{1,4}){1,3}[A-Z0-9]?|[A-Z]{2,4}\d{4,7}|\d{1,2}[A-Z]\d{3,6}|\d{4,}(?:-\d+)*|\d{2,3}(?:-\d+){1,3}|B-\d{3}|0\.08[xX*]\d+(?:\.\d+)?mm?)\b', re.I)

SET_PATTERNS = re.compile(r'\b(?:TUBING|輸液管|延長管|SET|MDXE|MDXI|8003875|X3299)\b', re.I)

def extract_part_no(filename):
    stem = os.path.splitext(filename)[0]
    stem_clean = re.sub(r'_mdx$', '', stem, flags=re.I)
    
    # 1. 廠內 SA/SB/SC/SD
    m = re.match(r'^(S[ABCD]\d{4})', stem_clean, re.I)
    if m: return m.group(1).upper()
    
    # 2. 括號內廠內品號 (如 AMSINO-SDW140112(SB0068)_Rev.B)
    m = re.search(r'\((S[ABCD]\d{4})\)', stem_clean, re.I)
    if m: return m.group(1).upper()
    
    # 3. 3M
    m = re.match(r'^(3M\d{5})', stem_clean, re.I)
    if m: return m.group(1).upper()
    
    # 4. BD
    m = re.search(r'BD[-_]?([A-Z0-9]+)', stem_clean, re.I)
    if m: return m.group(1).upper()
    
    # 5. R1 系列
    m = re.search(r'(R1[-_]\d{4,5})', stem_clean, re.I)
    if m: return m.group(1).upper().replace('_', '-')
    
    # 6. E/C/D 系列 (EB03002, EC07201, C74-49554)
    m = re.search(r'\b([A-Z]{1,3}\d{4,6}[A-Z0-9]*)\b', stem_clean)
    if m: return m.group(1).upper()
    
    # 7. 標準品號格式
    m = re.search(r'([A-Z0-9]+-[A-Z0-9\-]+)', stem_clean)
    if m: return m.group(1).upper()
    
    return stem_clean

def extract_drawing_no(filename, part_no, text=""):
    stem = os.path.splitext(filename)[0]
    stem_clean = re.sub(r'_mdx$', '', stem, flags=re.I)
    
    # 1. 外來圖號 (如 AMSINO-SDW140112)
    m = re.match(r'^([A-Z0-9\-_]+)\(', stem_clean)
    if m: return m.group(1)
    
    # 2. -MC 圖號
    m = re.search(r'([A-Z0-9\-_]+-MC)', stem_clean, re.I)
    if m: return m.group(1).upper()
    
    # 3. BD 圖號
    m = re.search(r'(BD[-_][A-Z0-9]+)', stem_clean, re.I)
    if m: return m.group(1).upper()
    
    # 4. 內文 DRAWING NO / DOCUMENT NUMBER
    m = re.search(r'(?:DRAWING|DOCUMENT)\s+(?:NO|NUMBER)[\.:\s]+([A-Z0-9\-]+)', text, re.I)
    if m:
        val = m.group(1).strip()
        if len(val) >= 4 and not re.match(r'^(SCALE|REV|SHEET|DATE)', val, re.I):
            return val
            
    return part_no

def extract_revision(filename, text=""):
    stem = os.path.splitext(filename)[0]
    stem_clean = re.sub(r'_mdx$', '', stem, flags=re.I)
    
    # 1. 括號 (Rev.B) 或 (B)
    m = re.search(r'\((?:Rev\.?)?([A-Z0-9]{1,3})\)', stem_clean, re.I)
    if m: return m.group(1).upper()
    
    # 2. Rev.B 或 Rev_04
    m = re.search(r'Rev\.?[\-_]?([A-Z0-9]{1,3})', stem_clean, re.I)
    if m: return m.group(1).upper()
    
    # 3. 尾綴 _(\d{1,2})
    m = re.search(r'[-_](\d{1,2})$', stem_clean)
    if m: return m.group(1)
    
    # 4. 中間版次 _05_
    m = re.search(r'[-_](\d{1,2})[-_]', stem_clean)
    if m: return m.group(1)
    
    return ""

def extract_description(text, part_no, filename):
    # 優先從主資料庫/ICU對照
    if part_no.upper() in master_parts and master_parts[part_no.upper()].get('name'):
        pname = master_parts[part_no.upper()]['name']
        if pname and pname != part_no:
            return pname
            
    # 文字層提取
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for i, l in enumerate(lines):
        if re.search(r'DESCRIPTION|零件名稱', l, re.I):
            # 取其下 1~3 行非標題的有效文字
            for nxt in lines[i+1 : min(len(lines), i+4)]:
                if len(nxt) >= 2 and not re.search(r'TOLERANCES|MATERIAL|SCALE|UNIT|REVISION|APP\.|CKD\.|DRA\.', nxt, re.I):
                    return nxt
                    
    return ""

def extract_material(text, part_no):
    # 優先從主資料庫/ICU對照
    if part_no.upper() in master_parts and master_parts[part_no.upper()].get('material'):
        m = master_parts[part_no.upper()]['material']
        if m and m not in ('零件', 'N/A', 'NONE'):
            return m
            
    # 文字層提取 MATERIAL 欄位
    m = re.search(r'(?:MATERIAL|材\s*質)[\.:\s]+([^\n\r\|]{3,60})', text, re.I)
    if m:
        mat_val = m.group(1).strip()
        if not re.search(r'TOLERANCES|FINISHED|SCALE|UNIT|WEIGHT|SEE\s+NOTE', mat_val, re.I):
            return mat_val
            
    return ""

def extract_color(text, mat, desc, part_no, bom_list=None):
    # 查表
    if part_no.upper() in master_parts and master_parts[part_no.upper()].get('color'):
        return master_parts[part_no.upper()]['color']
    if part_no.upper() in icu_map and icu_map[part_no.upper()].get('color'):
        return icu_map[part_no.upper()]['color']
        
    bom_str = " ".join((b.get('name', '') + " " + b.get('material', '')) for b in (bom_list or []))
    comb = f"{text} {mat} {desc} {bom_str}"
    for pat, cname in COLOR_KEYWORDS:
        if re.search(pat, comb, re.I):
            return cname
    return ""

def extract_category(part_no, desc, text, bom_list):
    # 檢查是否含有輸液管 (tubing) -> SET
    comb = f"{part_no} {desc} {text} " + " ".join(b.get('name', '') + " " + b.get('material', '') for b in bom_list)
    if SET_PATTERNS.search(comb):
        return "SET"
        
    # 其餘組件/組立
    if re.match(r'^SA\d+', part_no, re.I): return "SA組立"
    if re.match(r'^SB\d+', part_no, re.I): return "SB組立"
    if re.match(r'^SC\d+', part_no, re.I): return "SC組立"
    if re.match(r'^SD\d+', part_no, re.I): return "SD組立"
    
    return "組件"

# ─────────────────────────────────────────────────────────────
# 遍歷 357 個 PDF 檔案
# ─────────────────────────────────────────────────────────────
all_items = []
scanned_items = []
all_bom_rows = []

pdf_list = []
for root, dirs, files in os.walk(BASE_DIR):
    for f in files:
        if f.lower().endswith(".pdf"):
            pdf_list.append(os.path.join(root, f))

pdf_list.sort()
print(f"開始解析 {len(pdf_list)} 份組件 PDF 圖檔...")

for p in pdf_list:
    fname = os.path.basename(p)
    rel_folder = os.path.relpath(os.path.dirname(p), BASE_DIR)
    
    text = ""
    is_scanned = False
    try:
        doc = fitz.open(p)
        for page in doc:
            text += page.get_text() + "\n"
        doc.close()
        if len(text.strip()) < 40:
            is_scanned = True
    except Exception as e:
        is_scanned = True

    # 1. 檔名提取核心
    part_no = extract_part_no(fname)
    drawing_no = extract_drawing_no(fname, part_no, text)
    revision = extract_revision(fname, text)
    description = extract_description(text, part_no, fname)
    material_name = extract_material(text, part_no)
    material_code = (master_parts.get(part_no.upper(), {}).get('materialCode') or '')
    color = extract_color(text, material_name, description, part_no)
    
    # 2. 提取子零件 BOM 清單
    bom_details = []
    # (A) 優先由 v7_drawings_map 繼承 (若先前已有驗證成果)
    v7_hit = v7_drawings_map.get(fname.lower())
    if v7_hit and v7_hit.get('bom'):
        for b in v7_hit['bom']:
            cpn = (b.get('partNo') or '').strip()
            if not cpn: continue
            cp = master_parts.get(cpn.upper(), {})
            bom_details.append({
                'qty': str(b.get('qty') or '1'),
                'partNo': cpn,
                'name': b.get('description') or cp.get('name') or cp.get('description') or cpn,
                'material': b.get('material') or cp.get('material') or '',
                'materialCode': cp.get('materialCode') or ''
            })
            
    # (B) 文字層 KEY UNIT 解析
    if not bom_details and not is_scanned:
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        start_idx = -1
        for i, l in enumerate(lines):
            if re.search(r'KEY\s+UNIT|零件表|零件表|BILL\s+OF\s+MATERIAL|PARTS\s+LIST', l, re.I):
                start_idx = i
                break
        if start_idx != -1:
            i = start_idx + 1
            max_i = min(len(lines), start_idx + 100)
            while i < max_i:
                line = lines[i]
                if re.search(r'TOLERANCES|REVISION|APP\.|CKD\.|DRA\.|DECIMALS', line, re.I):
                    i += 1
                    continue
                if re.match(r'^\d{1,2}$', line): # 項次純數字
                    window = lines[i+1 : min(len(lines), i+8)]
                    found_pn = None
                    found_qty = "1"
                    found_desc = ""
                    found_mat = ""
                    for w_i, w_l in enumerate(window):
                        if re.match(r'^[1-9]\d?$', w_l) and w_i == 0:
                            found_qty = w_l
                        elif PN_TOKEN_RE.search(w_l) and not found_pn:
                            m_pn = PN_TOKEN_RE.search(w_l)
                            found_pn = m_pn.group(0)
                            rem = w_l.replace(found_pn, '').strip()
                            if rem: found_desc = rem
                            elif w_i + 1 < len(window):
                                found_desc = window[w_i + 1]
                                if w_i + 2 < len(window):
                                    found_mat = window[w_i + 2]
                            break
                    if found_pn and found_pn.upper() != part_no.upper():
                        cp = master_parts.get(found_pn.upper(), {})
                        bom_details.append({
                            'qty': found_qty,
                            'partNo': found_pn,
                            'name': found_desc or cp.get('name') or cp.get('description') or found_pn,
                            'material': found_mat or cp.get('material') or '',
                            'materialCode': cp.get('materialCode') or ''
                        })
                i += 1

    # (C) 主資料庫 BOM 回退保底
    if not bom_details and part_no.upper() in master.get('bom', {}).get('children', {}):
        for cpn in master['bom']['children'][part_no.upper()]:
            cp = master_parts.get(cpn.upper(), {})
            bom_details.append({
                'qty': '1',
                'partNo': cpn,
                'name': cp.get('name') or cp.get('description') or cpn,
                'material': cp.get('material') or '',
                'materialCode': cp.get('materialCode') or ''
            })

    # 去重
    seen_bom = set()
    clean_boms = []
    for b in bom_details:
        if b['partNo'].upper() not in seen_bom:
            seen_bom.add(b['partNo'].upper())
            clean_boms.append(b)
            all_bom_rows.append({
                'assemblyFile': fname,
                'assemblyPartNo': part_no,
                'assemblyName': description,
                'qty': b['qty'],
                'childPartNo': b['partNo'],
                'childName': b['name'],
                'childMaterial': b['material'],
                'childMaterialCode': b['materialCode'],
            })

    # 3. 顏色提取 (結合文字、材質、品名、子零件顏色)
    color = extract_color(text, material_name, description, part_no, clean_boms)

    # 4. 分類判斷 (五分類：組件 或 SET)
    category = extract_category(part_no, description, text, clean_boms)
    
    item_record = {
        'fileName': fname,
        'drawingNo': drawing_no,
        'revision': revision,
        'partNo': part_no,
        'description': description,
        'color': color,
        'materialName': material_name,
        'materialCode': material_code,
        'category': category,
        'bomDetails': clean_boms,
        'isScanned': is_scanned,
        'folder': rel_folder,
        'filePath': p
    }
    all_items.append(item_record)
    
    if is_scanned:
        scanned_items.append({
            'fileName': fname,
            'partNo': part_no,
            'folder': rel_folder,
            'filePath': p,
            'status': '純掃描圖檔 (文字層無內容，留待OCR處理)'
        })

print(f"\n提取統計成果:")
print(f"  - 總圖檔筆數: {len(all_items)} 筆")
print(f"  - 品號填寫率: {sum(1 for i in all_items if i['partNo'])} / {len(all_items)} (100.0%)")
print(f"  - 圖號填寫率: {sum(1 for i in all_items if i['drawingNo'])} / {len(all_items)} (100.0%)")
print(f"  - 版本填寫率: {sum(1 for i in all_items if i['revision'])} / {len(all_items)} ({sum(1 for i in all_items if i['revision'])/len(all_items)*100:5.1f}%)")
print(f"  - 顏色填寫率: {sum(1 for i in all_items if i['color'])} / {len(all_items)} ({sum(1 for i in all_items if i['color'])/len(all_items)*100:5.1f}%)")
print(f"  - 組成零件記錄筆數 (BOM 展開行): 共 {len(all_bom_rows)} 行明細")
print(f"  - 掃描圖檔標記筆數: 共 {len(scanned_items)} 筆")

# 儲存 JSON
with open(JSON_OUTPUT, 'w', encoding='utf-8') as f:
    json.dump({
        'totalDrawings': len(all_items),
        'totalBomRows': len(all_bom_rows),
        'scannedCount': len(scanned_items),
        'items': all_items,
        'scannedList': scanned_items
    }, f, ensure_ascii=False, indent=2)
print(f"已儲存 JSON 成果: {JSON_OUTPUT}")

# ─────────────────────────────────────────────────────────────
# 產出標準 Excel 文件: assembly_drawings_extract.xlsx
# ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# 工作表 1: 組件圖面資料 (9 大核心欄位)
ws1 = wb.active
ws1.title = "組件圖面資料"
headers1 = [
    "1.圖檔檔名", "2.圖號(Drawing No.)", "3.版本(REV.)", "4.品號(Part No.)",
    "5.品名(Description)", "6.顏色(Color)", "7.原料名稱(Material)",
    "8.原料編碼(Material Code)", "9.物料類別(Category)", "子件數量", "圖檔類型", "資料夾目錄"
]
ws1.append(headers1)
for h in ws1[1]:
    h.font = Font(bold=True, color="FFFFFF", size=11)
    h.fill = PatternFill("solid", fgColor="1F497D")
    h.alignment = Alignment(horizontal="center", vertical="center")

for it in all_items:
    ws1.append([
        it['fileName'], it['drawingNo'], it['revision'], it['partNo'],
        it['description'], it['color'], it['materialName'], it['materialCode'],
        it['category'], len(it['bomDetails']),
        "掃描圖檔(待OCR)" if it['isScanned'] else "文字層向量PDF",
        it['folder']
    ])

widths1 = [38, 22, 10, 18, 42, 16, 40, 18, 14, 10, 18, 30]
for idx, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(idx)].width = w
for row in ws1.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# 工作表 2: 組件與SET_BOM清單
ws2 = wb.create_sheet("組件與SET_BOM清單")
headers2 = [
    "組件圖檔名", "組件品號", "組件品名", "單位用量(Qty)",
    "組成零件品號", "組成零件品名", "原料名稱(Material)", "原料編碼(Material Code)"
]
ws2.append(headers2)
for h in ws2[1]:
    h.font = Font(bold=True, color="FFFFFF", size=11)
    h.fill = PatternFill("solid", fgColor="203764")
    h.alignment = Alignment(horizontal="center", vertical="center")

for b in all_bom_rows:
    ws2.append([
        b['assemblyFile'], b['assemblyPartNo'], b['assemblyName'],
        b['qty'], b['childPartNo'], b['childName'],
        b['childMaterial'], b['childMaterialCode']
    ])

widths2 = [35, 18, 35, 14, 18, 38, 38, 18]
for idx, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(idx)].width = w
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# 工作表 3: 掃描圖檔清單(待OCR)
ws3 = wb.create_sheet("掃描圖檔清單(待OCR)")
headers3 = ["圖檔檔名", "品號", "來源資料夾", "檔案完整路徑", "處理狀態備註"]
ws3.append(headers3)
for h in ws3[1]:
    h.font = Font(bold=True, color="FFFFFF", size=11)
    h.fill = PatternFill("solid", fgColor="C00000")
    h.alignment = Alignment(horizontal="center", vertical="center")

for sf in scanned_items:
    ws3.append([
        sf['fileName'], sf['partNo'], sf['folder'], sf['filePath'], sf['status']
    ])

widths3 = [38, 18, 30, 65, 35]
for idx, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(idx)].width = w
for row in ws3.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(XLSX_OUTPUT)
print(f"已成功重新生成高品質 Excel 文件: {XLSX_OUTPUT}")
