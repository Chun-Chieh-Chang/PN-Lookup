#!/usr/bin/env python3
"""
PN-Lookup 圖面資料提取工具 v7 — LLM 語意驅動版
策略：
  1. 快速路徑：檔名 regex + ICU lookup (秒級)
  2. 失敗案例：送 LLM (agnes-2.0-flash) 語意理解提取
  3. BOM：LLM 從完整文本語意提取
  4. 結果：JSON + Excel
"""
import json, re, glob, os, sys, time, subprocess, traceback
from pathlib import Path
from PIL import Image, ImageFilter, ImageEnhance
import pytesseract
pytesseract.pytesseract.tesseract_cmd = r'C:/Users/3kids/AppData/Roaming/TRAE SOLO/ModularData/ai-agent/vm/tools/app/tesseract/tesseract.exe'

# ── Configuration ───────────────────────────────────────────────────────────
BASE = Path(r"D:\Self-developed_Apps\PN-Lookup\rawdata\Drawings\零件")
OUT_XLSX = Path(r"D:\Self-developed_Apps\PN-Lookup\data\drawings_extract_v7.xlsx")
OUT_JSON = Path(r"D:\Self-developed_Apps\PN-Lookup\data\drawings_extract_v7.json")
OUT_SCAN = Path(r"D:\Self-developed_Apps\PN-Lookup\data\drawings_scanned_v7.json")
ICU_LOOKUP = Path(r"D:\Self-developed_Apps\PN-Lookup\data\icu-parts.json")

LLM_PROMPT_TEMPLATE = """Extract structured data from this engineering drawing text. Output pure JSON only (no markdown, no explanations).

Required fields:
- partNo: part number (format like R1-8392, E26-000-416, VLV-135-015, 75-2504)
- revision: revision (like 01, A, B, C)
- description: part name (English or Chinese, actual part name only — NOT REV/SHEET/SCALE/DIMENSIONS/CHECK/TOLERANCE/SURFACE TREATMENT text)
- materialName: material (like POLYPROPYLENE, ABS, POLYCARBONATE, PVC)
- color: color (like WHITE, BLACK, BLUE, CLEAR — omit if not found)
- materialCode: material code (like COMMODITY #AB002)

Rules:
1. Do NOT extract description from REVISION HISTORY table
2. Do NOT extract description from SCALE/DIMENSIONS/tolerance text
3. Do NOT extract description from CHECK POINT/ACCEPTANCE CRITERIA
4. description must be the actual part name/function (like BREATHER CAP, VALVE BODY)
5. partNo should come from DRAWING # or PART NO., NOT from COMMODITY #
6. Leave unfound fields as empty string ""

Drawing text:
{text}
"""

BOM_PROMPT_TEMPLATE = """Extract all assembly parts from this engineering drawing BOM. Output a JSON array only (no markdown, no explanations).

Each item must have:
- partNo: part number (like R1-3152)
- description: part name
- material: material (like POLYPROPYLENE, BLUE)
- qty: quantity (number, empty string if not found)

Rules:
1. Only extract items that have a part number
2. Exclude tools, gauges, purchased parts unless explicitly listed as BOM items
3. Note Markdown table rows (format like | ... PART NAME(PART NO) | ... |)

Drawing text:
{text}
"""

# ── Helpers ─────────────────────────────────────────────────────────────────
def path_to_posix(p):
    return str(p).replace('\\', '/').lower()

def get_folder_type(filepath):
    p = path_to_posix(filepath)
    if 'icu' in p and '含量測點' in p: return 'icu_meas'
    if 'icu' in p and '原始檔' in p:   return 'icu_raw'
    if 'icu' in p:                    return 'icu'
    if '廠內零件圖面' in p:            return 'factory'
    if '/part/' in p:                 return 'part'
    if '/tubing' in p:                return 'tubing'
    if '綜合圖面' in p:               return 'comprehensive'
    return 'other'

def parse_llm_text(text_out):
    """Parse LLM response text — handles JSON inline or markdown code blocks."""
    if not text_out:
        return None
    text_out = text_out.strip()

    # Try direct JSON first
    try:
        obj = json.loads(text_out)
        if isinstance(obj, dict):
            return obj
        if isinstance(obj, list):
            return obj
    except Exception:
        pass

    # Try JSON in backticks
    m = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', text_out)
    if m:
        try:
            obj = json.loads(m.group(1))
            if isinstance(obj, dict):
                return obj
            if isinstance(obj, list):
                return obj
        except Exception:
            pass

    # Try find first { in text
    idx = text_out.find('{')
    if idx >= 0:
        try:
            end_idx = text_out.rfind('}')
            if end_idx > idx:
                obj = json.loads(text_out[idx:end_idx+1])
                if isinstance(obj, dict):
                    return obj
        except Exception:
            pass

    return None

# ── Material extraction from PDF text ───────────────────────────────────────
MAT_KEYWORDS = [
    # Thermoplastics
    'POLY', 'ABS', 'PVC', 'RESIN', 'NYLON', 'PEEK', 'TPU', 'TPE',
    'PP', 'PS', 'PC', 'PMMA', 'POLYCARB', 'POLYPROP', 'POLYETHYLENE',
    'HDPE', 'LDPE', 'POM', 'PTFE',
    # Elastomers
    'SILICONE', 'RUBBER', 'EPDM', 'NBR', 'VITON', 'FLUOROCARBON',
    # Metals (full names only — avoid short substrings like AL, SS)
    'BRASS', 'COPPER', 'STEEL', 'STAINLESS', 'SUS304', 'SUS316', 'SUS',
    'ALUMINUM', 'ALUMINIUM', 'TITANIUM', 'TITAN',
    # Others
    'ULTEM', 'ADEKA',
]
# Material aliases used in titles — single-line keyword must match one of these
MAT_ALIASES = [
    'HDPE', 'PP', 'POM', 'PTFE', 'NYLON', 'PC', 'ABS', 'PVC', 'BRASS',
    'COPPER', 'STEEL', 'SUS', 'ALUMINUM', 'SILICONE', 'RUBBER', 'EPDM', 'NBR',
    'PTFE-COTTON', 'POLYPROPYLENE', 'POLYETHYLENE', 'POLYCARBONATE',
]
# Aliases eligible for fuzzy matching — must be >= 4 chars to avoid false positives
FUZZY_ALIASES = [
    'HDPE', 'POM', 'PTFE', 'NYLON', 'ABS', 'PVC',
    'BRASS', 'STEEL', 'SUS', 'SILICONE', 'RUBBER', 'EPDM', 'NBR',
    'PTFE-COTTON', 'POLYPROPYLENE', 'POLYETHYLENE', 'POLYCARBONATE',
]
# Aliases whose fuzzy matches require >= 80% threshold (short aliases)
FUZZY_SHORT = {'PP', 'PC', 'PS'}
# Lines to skip — these are label/header lines, not material values
MAT_SKIP_LINES = {'MATERIAL', '材質', '材  質', '材　質', '表面處理', 'FINISHED',
                  '重   量', 'WEIGHT', '比  例', 'SCALE', '單位', 'UNIT',
                  '核  准', 'APP.', '核  對', 'C.K.D.', '製  圖', 'DRA.',
                  '簽    名', 'BY', '日    期', 'DATE', '尺寸級別',
                  'DIM. CLASSIFICATION', '重要尺寸', 'CRITICAL', '次要尺寸', 'MAJOR',
                  '一般尺寸', 'ALL OTHERS MINOR', 'DECIMALS', 'FRACTIONAL', 'ANGLES'}
MAT_NOISE = ['SUPPLIER', 'PERFORM', 'INSPECTION', 'CERTIFICATION', 'TRACEABILITY', 'REQUIRE',
             'NOTE', 'DRAWING', 'SPECIFICATION', 'ACCEPTANCE', 'CRITERIA', 'SURFACE', '處理',
             'VISUAL', 'INCOMING', 'ENCAPSULATIONS', 'APPLICATION', 'APP.', 'INFO',
             'PRODUCT#', 'PLCS', 'DIMENSION', 'MEASURE', 'CHECK', 'REVISION', 'SCALE',
             'SHEET', 'SIZE', 'INTERPRET', 'THIRD ANGLE', 'UNLESS OTHERWISE',
             'CONDITIONALLY', 'CONFIDENTIAL', 'PROPRIETARY', 'REPRODUCE', 'DISCLOSE',
             'CONTENTS', 'HEREOF', 'MATERIAL CERTIFICATION', 'ENCAPSULATED', 'UPDATED',
             'STEPPED', 'CAPPED', 'VENT FILTER', 'STRAP', 'PERFORM',
             'TOLERANCES', 'GENERAL TOLERANCE', 'DECIMALS', 'FRACTIONAL', 'ANGLES',
             'DIM. CLASSIFICATION', '重要尺寸', '次要尺寸', '一般尺寸',
             'NEW DRAWING', 'REVISED', 'REVISION HISTORY', 'CHANGE',
             'FILE NO', 'DRAWING NO', 'PART NO', 'PART NUMBER', 'PRODUCT',
             # Legal/disclaimer terms (often appear in OCR text blocks)
             'POSSESSION', 'CONFERS', 'RIGHTS', 'TRANSFER', 'LICENSE', 'AGREEMENT',
             'MOULDEX', 'CO., LTD', 'LTD', 'INC', 'LLC',
             'CONFIDENTIAL', 'PROPERTY OF',
             # Chinese title block fields (not material values)
             '表面處理', '處理', '重量', '重量', '比例', '單位',
             '核准', '核對', '製圖', '簽名', '日期', '確認', '查對',
             # Common OCR-garbled instructional/noise phrases
             'CLOSED', 'OPEN AND', 'APPROVAL', 'DEFINITION', 'VESSEL',
             'LEVEL', 'RIGHT SIDE', 'UNACCEPTABLE']
MAT_WEAK = ['APP.', 'INFO', 'NOTE', 'SEE', 'PER', 'ITEM', 'PART', 'NO.']
# Single-word material abbreviations that are valid even without separators
MAT_SINGLE_WORD_ABBREVS = {
    'ABS', 'PC', 'PP', 'PE', 'PVC', 'PS', 'PMMA', 'POM', 'PTFE', 'PU',
    'TPU', 'TPE', 'NYLON', 'HDPE', 'LDPE', 'BRASS', 'COPPER', 'STEEL',
    'SUS', 'EPDM', 'NBR', 'VITON', 'SILICONE', 'RUBBER',
}

def _is_material_line(line):
    """Check if a stripped line looks like a material specification."""
    lu = line.upper().strip()
    if not lu or len(lu) < 2 or len(lu) > 200:
        return False
    if any(n in lu for n in MAT_NOISE):
        return False
    if any(w == lu for w in MAT_WEAK):
        return False
    return True

def _has_material_keyword(line):
    """Check if line contains any known material keyword."""
    lu = line.upper()
    return any(k in lu for k in MAT_KEYWORDS)

def _material_starts_valid(mat_name):
    """Validate that a material name is plausible — starts with a material keyword/abbrev."""
    lu = mat_name.strip().upper()
    # Must start with a material keyword, single-word abbrev, or descriptor (OPAQUE, WHITE, etc.)
    if not lu:
        return False
    first_word = lu.split()[0] if lu.split() else ''
    # Direct match: starts with keyword or abbreviation
    if any(k == first_word or first_word.startswith(k) for k in MAT_SINGLE_WORD_ABBREVS):
        return True
    if any(first_word.startswith(k) for k in MAT_KEYWORDS if len(k) >= 3):
        return True
    # Descriptor words are ok: WHITE, BLACK, CLEAR, OPAQUE, TRANSPARENT
    DESCRIPTORS = {'WHITE', 'BLACK', 'CLEAR', 'OPAQUE', 'TRANSPARENT', 'BLUE', 'RED', 'GREEN',
                   'YELLOW', 'NATURAL', 'COLOR', 'COLOUR', 'HIGH', 'MEDICAL', 'GRASDE',
                   'GRADE', 'FOOD', 'LEVEL', 'RADIATION', 'STERILE', 'GAMMA'}
    if first_word in DESCRIPTORS:
        return True
    # Reject: lines starting with non-material words (e.g. "Item Number: SPC...")
    NON_MAT_STARTERS = {'ITEM', 'NUMBER', 'PERFORM', 'VISUAL', '物理', 'REVISION', 'CHANGE',
                        'REMOVE', 'COMPLETE', 'SUBMIT', 'FOLLOW', 'SPECIFY', 'APPLY',
                        'PART', 'ASSY', 'ASM', 'SECTION', 'NOTE', 'REQ', 'REQUIRE'}
    if first_word in NON_MAT_STARTERS:
        return False
    return True

def _fuzzy_material_match(line):
    """For OCR-garbled text: check if line resembles a material name even with typos.
    E.g. 'PVLENE' → matches POLYPROPYLENE via subsequence matching.
    Requirements:
      - line must be >= 4 chars
      - alias must be >= 4 chars (rejects 2-letter aliases for fuzzy)
      - ALL chars of alias must appear in order in the line (subsequence match)
      - line must not contain description keywords or name patterns"""
    lu = line.upper().strip()
    if len(lu) < 4:
        return False
    if not _is_material_line(lu):
        return False
    # Skip lines that look like part descriptions
    DESC_WORDS = {'CAP', 'VALVE', 'SPIKE', 'FILTER', 'CONNECTOR', 'ADAPTOR',
                  'ADAPTER', 'BODY', 'RING', 'HUB', 'VENT', 'NOZZLE', 'TIP',
                  'HOLDER', 'CLAMP', 'COLLAR', 'BUSHING', 'SEAT', 'STEM',
                  'DUCTILE', 'NUT', 'BOLT', 'SCREW', 'PIN', 'SHAFT', 'GEAR',
                  'BEARING', 'WASHER', 'PLATE', 'DISC', 'DISK', 'TUBE', 'HOSE',
                  'FITTING', 'SLEEVE', 'EXTENDER', 'INSERT', 'SEAL', 'MEMBRANE',
                  'RETAINER', 'BREATHER', 'RESTRICTOR', 'REGULATOR',
                  'NEW DRAWING', 'REVISED', 'REVISION', 'CHANGE', 'UPDATED',
                  'FILE NO', 'PART NO', 'DRAWING NO', 'MOULDEX', 'CO., LTD',
                  'CONFIDENTIAL', 'PROPERTY OF', 'DCC CONTROL',
                  'LUER', 'THREAD', 'TAPER', 'GAGE', 'GAUGE', 'ISO',
                  'FROM', 'TO', 'PERFORM', 'INSPECTION', 'CERTIFICATION'}
    if any(dw in lu for dw in DESC_WORDS):
        return False
    # Must not match common name patterns (two capitalized words = person name)
    if re.match(r'^[A-Z][a-z]+ [A-Z][a-z]+$', lu.strip()):
        return False
    # Subsequence match: ALL chars of alias must appear in order in the line
    for alias in FUZZY_ALIASES:
        if len(alias) < 4:
            continue
        alias_up = alias.upper()
        j = 0
        for c in lu:
            if j < len(alias_up) and c == alias_up[j]:
                j += 1
        if j == len(alias_up):  # ALL chars matched in order
            return True
    return False

def _best_material_candidate(candidates):
    """Given a list of candidate strings, return the best one.
    Preference: longer, more keyword hits, no revision-note noise."""
    REV_NOISE = ['REVISED', 'REMOVED', 'CHANGED', 'UPDATED', 'NOTE', 'NOTES',
                 'REVISION', 'KEEP', 'KEEPED', 'AND', 'FROM', 'TO', 'BY']
    def score(line):
        s = len(line)
        kw_hits = sum(1 for k in MAT_KEYWORDS if k in line.upper())
        s += kw_hits * 10
        if any(n in line.upper() for n in REV_NOISE):
            s -= 100  # strongly penalise revision-note noise
        # Penalise verbose disclaimer/legal text (many words, no keyword hit)
        word_count = len(line.split())
        if word_count > 10 and kw_hits == 0:
            s -= 80
        return s
    return max(candidates, key=score) if candidates else ''

def _looks_like_material_value(line):
    """Strict validation: does this line actually look like a material specification?"""
    # Reuse _is_material_line to catch MAT_NOISE words consistently
    if not _is_material_line(line):
        return False
    # Strip non-ASCII (OCR artifacts like Chinese glyphs) so "ABS T-900 国" → "ABS T-900" validates
    clean_line = re.sub(r'[^\x00-\x7F]', ' ', line).strip()
    lu = clean_line.upper()
    # Reject pipe-heavy OCR garbage (e.g. "N | Ppesst | ] 083g | 21 |")
    pipe_count = lu.count('|')
    non_space_len = len(lu.replace(' ', ''))
    if pipe_count >= 3 and non_space_len > 0 and pipe_count / non_space_len > 0.1:
        return False
    # Allow known single-word material abbreviations FIRST (before length checks)
    if lu in MAT_SINGLE_WORD_ABBREVS:
        return True
    if len(lu) < 3 or len(lu) > 120:
        return False
    # Reject single-word short labels (APP., NOTE., etc.)
    if ' ' not in lu and len(lu) < 6:
        return False
    # Reject lines starting with digit + period (revision notes: "1. ", "2. ")
    # BUT allow single-digit list items like "1. Material: ABS..." which are valid
    if re.match(r'^\d+\.\s', lu) and not re.match(r'^\d\.\s+[A-Z]', lu):
        return False
    # Reject lines starting with single letter + period (e.g., "A. ", "C.")
    if re.match(r'^[A-Z]\.$', lu) or re.match(r'^[A-Z]\.\s', lu):
        return False
    # Reject address-like patterns (number + street name)
    if re.match(r'^\d+\s+[A-Z][a-z]+', lu):
        return False
    # Reject person name patterns: "INITIALS.NAME" or "FIRST LAST" (short)
    if re.match(r'^[A-Z]\.[A-Z]+$', lu):
        return False
    if re.match(r'^[A-Z][a-z]+\s+[A-Z][a-z]+$', line.strip()) and len(lu) < 25:
        return False
    # Reject person names with initials: "Erwin L.", "Chen W."
    if re.match(r'^[A-Z][a-z]+\s+[A-Z]\.', line.strip()):
        return False
    # Reject lines starting with part number style
    if re.match(r'^[A-Z]{1,4}-\d', lu):
        return False
    # Reject lines that are entirely part-number lists (e.g., "R1-10134, R1-15853")
    tokens = [t.strip() for t in lu.split(',') if t.strip()]
    if tokens and all(re.match(r'^[A-Z]\d-[A-Z0-9]+$', t) for t in tokens):
        return False
    # Reject revision note patterns (contains "FROM ... TO" structure)
    if ' FROM ' in lu and ' TO ' in lu:
        return False
    # Reject revision note continuation lines: "to \"ABS...\"" or "From \"ABS...\""
    if re.match(r'^(?:TO|FROM)\s*["\"]?', lu):
        return False
    # Reject company name suffixes: "MOULDEX CO., LTD."
    if re.search(r'\b(?:CO\.?|LTD\.?|INC\.?|LLC)\s*$', lu):
        return False
    # Reject lines starting with common English words (not material names)
    ENGLISH_STARTERS = {'WE', 'THE', 'THIS', 'THAT', 'TO', 'AND', 'BUT', 'FOR', 'WITH',
                        'ARE', 'HAS', 'HAD', 'HAVE', 'CAN', 'WAS', 'WERE', 'BEEN', 'BEING',
                        'DOES', 'DONE', 'SHOULD', 'COULD', 'WOULD', 'MUST', 'NEED', 'WILL',
                        'WANT', 'ONLY', 'JUST', 'LIKE', 'ABOUT', 'DOWN', 'OVER', 'BACK',
                        'TURN', 'MAKE', 'TAKE', 'COME', 'GET', 'GIVE', 'USE', 'CALLED',
                        'REQUEST', 'SIGNATURE', 'IN', 'FROM', 'PRODUCT'}
    first_word = lu.split()[0] if lu.split() else ''
    if first_word in ENGLISH_STARTERS:
        return False
    # Reject lines containing dates (e.g., "Jun 18, 2019" or "Jan. 3, 2023")
    if re.search(r'\b(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[. ]\s*\d', lu):
        return False
    # Reject lines with non-ASCII characters (already stripped above; this is a safety net)
    if not re.match(r'^[\x00-\x7F ]+$', clean_line):
        return False
    # Must start with material abbreviation or uppercase word followed by space/comma/dash
    # Allow single-digit list items like "1. Material: ..." since they appear in engineering docs
    # Expanded prefix check: {1,12} to allow descriptor words like OPAQUE, TRANSPARENT, CLEAR
    if re.match(r'^\d\.\s+(?:MATERIAL|材質)', lu):
        return True
    if re.match(r'^[A-Z]{1,12}[\s,.\-/]', lu):
        return True
    # Allow part-number-style prefixes: "R1-1034, ..."
    if re.match(r'^[A-Z]\d', lu):
        return True
    # All-caps material word >= 4 chars (e.g., POLYPROPYLENE, POLYCARBONATE)
    if re.match(r'^[A-Z][A-Z]+$', lu) and len(lu) >= 4:
        return True
    # Resin grade codes: 2-4 uppercase letters + 3+ digits (+ optional trailing letters), e.g. PP6331, HD810MO, GP2244
    if re.match(r'^[A-Z]{2,4}\d{3,}[A-Z0-9]*$', lu):
        return True
    return False

def extract_material_from_text(text):
    """Extract material name and code from PDF text using multiple strategies."""
    mat_name = ''
    mat_code = ''

    # Pattern 1: Look backward from MATERIALS: label (ICU-style spec sheets)
    idx_mat = text.upper().find('MATERIALS:')
    if idx_mat >= 0:
        search_area = text[max(0, idx_mat - 500):idx_mat]
        candidates = []
        for line in reversed(search_area.split('\n')):
            line = line.strip().rstrip(',')
            if _has_material_keyword(line) and _is_material_line(line):
                candidates.append(line)
        if candidates:
            mat_name = _best_material_candidate(candidates)

    # Pattern 2: 'MATERIAL:' or 'Material:' label — scan forward for value lines
    if not mat_name:
        for m in re.finditer(r'(?:MATERIAL|Material)\s*:\s*$', text, re.MULTILINE):
            start = m.end()
            # Collect subsequent non-empty lines until we hit a new section
            chunk = text[start:start+500]
            for line in chunk.split('\n'):
                line = line.strip().rstrip(',')
                if not line:
                    continue
                lu = line.upper()
                # Stop at section headers
                if lu in ('MATERIAL CERTIFICATION', 'MATERIAL NOTES', 'MATERIALS:',
                          'SUPPLIERS INFO', 'SUPPLIER', 'PRODUCT#', 'DOCUMENT #'):
                    break
                if any(kw in lu for kw in ['PERFORM', 'INSPECTION', 'CERTIFICATION',
                                            'TRACEABILITY', 'REQUIRE', 'SPECIFICATION']):
                    continue
                # Reject noise patterns:
                # - Single letter + period (e.g., "A. INSPECT...", "C.")
                # - Digit + period that is NOT a material label line (e.g., "1. Material: ...")
                if re.match(r'^[A-Z]\.\s', lu):
                    continue
                if re.match(r'^\d+\.\s', lu) and not re.match(r'^\d+\.\s+(?:MATERIAL|Material)\b', lu, re.IGNORECASE):
                    continue
                # - Person name pattern (two capitalized words with space)
                if re.match(r'^[A-Z][a-z]+\s+[A-Z][a-z]+$', lu.strip()):
                    continue
                # - Dimension/coordinate patterns (±, inches, etc.)
                if any(c in line for c in ['±', 'INCH', 'MM', 'GAUGE FIG', 'ISO 594', 'ISO 80369']):
                    continue
                # - Lines starting with part number
                if re.match(r'^[A-Z]{1,3}-\d+', lu):
                    continue
                # Must contain a material keyword or pass fuzzy check
                if _has_material_keyword(line) and _is_material_line(line) and _looks_like_material_value(line):
                    mat_name = line
                    break
                if _fuzzy_material_match(line) and _is_material_line(line) and _looks_like_material_value(line):
                    mat_name = line
                    break
            if mat_name:
                break

    # Pattern 2b: 'N. Material: value' — value on same line (e.g. VLV drawings)
    if not mat_name:
        for line in text.split('\n'):
            m = re.search(r'^\d+\.\s+Material\s*:\s*(.+)$', line, re.IGNORECASE)
            if m:
                candidate = m.group(1).strip().rstrip('.').strip()
                if candidate and len(candidate) >= 2:
                    mat_name = candidate
                    break

    # Pattern 3: COMMODITY NO./# → material code
    m_comm = re.search(r'COMMODITY\s*(?:NO\.|#)\s*([A-Z]?\d+-\d+)', text, re.IGNORECASE)
    if m_comm:
        mat_code = m_comm.group(1)

    # Pattern 4: Three-line title block — 材質 / MATERIAL / VALUE
    # Common in Chinese engineering drawings: lines like
    #   '材   質'        (or 'MATERIAL')
    #   'MATERIAL'       (or '材質')
    #   ... various label lines ...
    #   'ABS TOYOLAC 900'  ← actual value (typically within 15-25 lines)
    if not mat_name:
        lines = text.split('\n')
        for i, line in enumerate(lines):
            lu = line.upper().strip()
            if lu in ('MATERIAL', '材質', '材  質', '材　質'):
                # Scan forward up to 25 lines for the material value
                # Skip label rows (Chinese or English field names)
                LABEL_PATTERNS = {
                    '材質', '材  質', '材　質', '表面處理', '處理', 'FINISH',
                    '重量', 'WEIGHT', '比例', 'SCALE', '單位', 'UNIT',
                    '核准', '核對', '製圖', 'DRAWN', '簽名', 'BY', '日期', 'DATE',
                    '確認', 'APP.', 'C.K.D.', 'DRA.', 'FILE NO', 'PART NO',
                    '零件名稱', 'DESCRIPTION', 'REV', 'PAGE', '尺寸級別',
                    'DIM. CLASSIFICATION', '一般公差', 'TOLERANCES', 'DECIMALS',
                    'FRACTIONAL', 'ANGLES', 'SECTION', 'MM', 'INCH',
                    # Revision history patterns (must skip these)
                    'REVISION', 'REVISION HISTORY', 'REMOVED', 'UPDATED',
                    'NEW DRAWING', 'REVISED', 'CHANGE', 'CHANGED',
                    'FROM', 'TO', 'ADD', 'ADDED', 'REMOVE', 'REMOVED',
                }
                for j in range(i+1, min(i+25, len(lines))):
                    candidate = lines[j].strip().rstrip(',')
                    if not candidate:
                        continue
                    cu = candidate.upper()
                    if cu in LABEL_PATTERNS or cu in MAT_SKIP_LINES:
                        continue
                    # Must contain a material keyword or pass fuzzy check
                    if _has_material_keyword(candidate) or _fuzzy_material_match(candidate):
                        if _is_material_line(candidate) and _looks_like_material_value(candidate):
                            mat_name = candidate
                            break
                    # Also accept pure resin grade codes (e.g. PP6331, HD810MO, GP2244)
                    if re.match(r'^[A-Z]{2,4}\d{3,}[A-Z0-9]*$', candidate.upper()):
                        mat_name = candidate
                        break
                if mat_name:
                    break

    # Pattern 5: Best candidate from full-text scan (last resort, comprehensive)
    if not mat_name:
        candidates = []
        for line in text.split('\n'):
            line = line.strip()
            if (_has_material_keyword(line) or _fuzzy_material_match(line)) \
                    and _is_material_line(line) and _looks_like_material_value(line):
                candidates.append(line)
        if candidates:
            mat_name = _best_material_candidate(candidates)

    # Pattern 6: Fuzzy-only fallback (heavily OCR-garbled scanned PDFs)
    # Only accept short lines (≤6 words) AND requiring at least one keyword hit
    # to avoid false positives from legal/disclaimer text
    if not mat_name:
        for line in text.split('\n'):
            line = line.strip()
            if len(line.split()) > 6:
                continue
            if _fuzzy_material_match(line) and _is_material_line(line) and _looks_like_material_value(line):
                # Require at least one material keyword hit for fuzzy matches
                # (prevents "STEEL" subsequence from matching "Serving The Health Care Industry")
                if _has_material_keyword(line):
                    mat_name = line
                    break

    # Pattern 7: Title block table format — MATERIAL header with value in adjacent cell
    # PDF text extraction reads tables top-to-bottom, left-to-right, so headers and
    # values appear on different lines. We find the MATERIAL header and scan forward.
    if not mat_name:
        lines = text.split('\n')
        for i, line in enumerate(lines):
            lu = line.strip().upper()
            # Match MATERIAL header (with possible Chinese spacing)
            if re.match(r'^(?:MATERIAL|材\s*質|材　質)$', lu) and len(lu) <= 15:
                # Look ahead up to 50 lines for a material value
                # Skip header-like lines (表面处理, FINISHED, WEIGHT, SCALE, UNIT, etc.)
                skip_headers = {
                    '表面處理', '處理', 'FINISH', 'FINISHED', '重', 'WEIGHT',
                    '比  例', 'SCALE', '單位', 'UNIT', '核  准', 'APP.', '核  對',
                    'C.K.D.', 'C. K. D.', '製  圖', 'DRA.', '簽', 'BY', '日  期',
                    'DATE', 'DECIMALS/MM', 'FRACTIONAL', 'ANGLES', 'DIM',
                    'CLASSIFICATION', 'CRITICAL', 'MAJOR', 'MINOR', 'REV',
                    'REVISION', '確認', 'CKD', '文件編號', 'FILE NO.', 'DRAWING',
                    'DOCUMENT', 'PART NO.', '零件編號', '零件名稱', 'DESCRIPTION',
                    'SIZE', 'THIRD ANGLE', 'TOLERANCES', 'DIM. CLASSIFICATION',
                    '重要尺寸', '次要尺寸', '一般尺寸', 'ALL OTHERS MINOR',
                    'GENERAL TOLERANCE', 'UNLESS OTHERWISE', 'DCC CONTROL',
                    'MOULDEX', 'ICU MEDICAL', 'NEW DRAWING', 'PAGE',
                }
                for j in range(i + 1, min(i + 50, len(lines))):
                    nl = lines[j].strip()
                    nlu = nl.upper()
                    if not nl:
                        continue
                    # Stop at noise sections
                    if nlu in skip_headers or any(h in nlu for h in skip_headers):
                        continue
                    if any(skip in nlu for skip in ['PERFORM', 'INSPECTION', 'CERTIFICATION',
                                                     'TRACEABILITY', 'REPRODUCE', 'DISCLOSE',
                                                     'CONFIDENTIAL', 'PROPRIETARY', 'CONDITIONALLY',
                                                     'ENCAPSULATION', 'INCOMING', 'SPECIFICATION']):
                        continue
                    # Reject dimension/coordinate lines
                    if re.match(r'^[-.\d]+\s*$', nlu) or '±' in nl or 'MM' in nlu or 'INCH' in nlu:
                        continue
                    # Check for material keywords
                    if _has_material_keyword(nl) and _is_material_line(nl) and _looks_like_material_value(nl):
                        mat_name = nl
                        break
                    # Check for material-like patterns even without keyword
                    # e.g. "PP6331", "HD810MO", "RX1805-451118", "GP2244"
                    if re.match(r'^[A-Z]{2,3}\d{4,}$', nlu) and len(nlu) <= 12:
                        # Likely a resin grade code (PP6331, HD810MO, RX1805)
                        if _looks_like_material_value(nl):
                            mat_name = nl
                            break
                    if re.match(r'^[A-Z]{2,4}-\d+$', nlu) and len(nlu) <= 15:
                        if _looks_like_material_value(nl):
                            mat_name = nl
                            break
                if mat_name:
                    break

    # Pattern 8: Revision note / change history — "Material: XXX" or "材料: XXX"
    if not mat_name:
        for line in text.split('\n'):
            line = line.strip()
            lu = line.upper()
            # Match "Material: XXX" or "材料: XXX" or "1.Material change from X to Y"
            m = re.match(r'^(?:.*(?:MATERIAL|材料|材質)\s*[:：]\s*)+(.+)$', lu)
            if m:
                value = m.group(1).strip().strip('"').strip("'").strip('.')
                if value and len(value) >= 2 and _is_material_line(value) and _looks_like_material_value(value):
                    mat_name = value
                    break
            # Match "from X to Y" pattern (material change note)
            m2 = re.search(r'(?:FROM|原)\s+[^T]*(?:TO|更改為|更改到|更改为)\s+(.+?)(?:\.|$)', lu)
            if m2:
                value = m2.group(1).strip().strip('"').strip("'")
                if value and len(value) >= 2 and _is_material_line(value) and _looks_like_material_value(value):
                    mat_name = value
                    break

    return mat_name, mat_code

# ── Description extraction from PDF text ────────────────────────────────────
DESC_KEYWORDS = [
    'CAP', 'VALVE', 'SPIKE', 'FILTER', 'CONNECTOR', 'ADAPTOR', 'ADAPTER',
    'BODY', 'RING', 'HUB', 'VENT', 'NOZZLE', 'TIP', 'COUPLER', 'SEAT',
    'DIAPHRAGM', 'LEAF', 'PLUNGER', 'STEM', 'BOOT', 'GROMMET', 'PLATE',
    'HOLDER', 'CONE', 'SLEEVE', 'CHECK', 'LUMEN', 'EXTENDER', 'INSERT',
    'SEAL', 'MEMBRANE', 'RETAINER', 'CLAMP', 'COLLAR', 'BUSHING',
    'BREATHER', 'EXTENSION', 'RESTRICTOR', 'REGULATOR', 'TRANSITION',
    'DISPENSING', 'INFUSION', 'CATETER', 'CATHETER',
]
DESC_NOISE = [
    'PERFORM', 'INSPECTION', 'CERTIFICATION', 'TRACEABILITY', 'REQUIRE',
    'NOTE', 'DRAWING', 'SPECIFICATION', 'ACCEPTANCE', 'CRITERIA', 'SURFACE',
    'VISUAL', 'INCOMING', 'ENCAPSULATIONS', 'APPLICATION', 'DIMENSION',
    'MEASURE', 'CHECK', 'REVISION', 'SCALE', 'SHEET', 'SIZE', 'INTERPRET',
    'THIRD ANGLE', 'UNLESS OTHERWISE', 'CONDITIONALLY', 'CONFIDENTIAL',
    'PROPRIETARY', 'REPRODUCE', 'DISCLOSE', 'CONTENTS', 'HEREOF',
    'MATERIAL CERTIFICATION', 'POLYPROPYLENE', 'POLYCARBONATE', 'COMMODITY',
    'ISO', 'DOCUMENT', 'ICU MEDICAL', 'MANUFACTURER', 'EFFECTIVE DATE',
]
DESC_PN_PATTERN = re.compile(r'(?:R1-\d+|[A-Z]\d+-\d+|\d{2,5}-\d{2,5})')

# ── Tesseract OCR fallback for scanned PDFs ──────────────────────────────────
TESSERACT_CMD = r'C:/Users/3kids/AppData/Roaming/TRAE SOLO/ModularData/ai-agent/vm/tools/app/tesseract/tesseract.exe'
OCR_TEXT_THRESHOLD = 500  # if extracted text < this AND no material found, try OCR


def _quality_filter(mat_name):
    """Reject garbage OCR results — keep only plausible material names."""
    if not mat_name:
        return None
    lu = mat_name.upper().strip()
    if '|' in mat_name or len(lu) > 60:
        return None
    # Grade codes: PP6331, HD810MO, GP2244
    if re.match(r'^[A-Z]{2,4}\d{3,}[A-Z0-9-]*$', lu):
        return mat_name
    # Short codes with digits: PP-6331, PC-1234
    if len(lu) <= 20 and re.match(r'^[A-Z]{2,6}[\d\-/]', lu):
        return mat_name
    # Full material names
    MAT_STARTS = ['POLYPROPYLENE','POLYCARBONATE','POLYETHYLENE','HIGH DENSITY',
                   'LOW DENSITY','POLYVINYL CHLORIDE','NYLON','PTFE','POM','TPE',
                   'TPU','EPDM','NBR','SILICONE','RUBBER','BRASS','COPPER',
                   'STEEL','SUS','ABS','POLYURETHANE','POLYAMIDE','POLYOLEFIN',
                   'POLYSULFONE','PEEK','ULTEM']
    if any(lu.startswith(kw) for kw in MAT_STARTS):
        return mat_name
    # Descriptor + material: WHITE POLYCARBONATE, BLACK NYLON 6/6
    DESCRIPTORS = {'WHITE','BLACK','CLEAR','OPAQUE','TRANSPARENT','BLUE','RED',
                   'GREEN','YELLOW','NATURAL','MEDICAL','COLOR','COLOUR'}
    words = lu.split()
    if words and words[0] in DESCRIPTORS and len(words) >= 2:
        second = words[1]
        if any(kw in second for kw in ['POLY','ABS','NYL','PTFE','POM','TPE','TPU',
                                        'EPDM','NBR','SIL','RUBBER','STEEL','BRASS','COPPER']):
            return mat_name
    if lu in MAT_SINGLE_WORD_ABBREVS:
        return mat_name
    return None


def _ocr_extract_material(filepath):
    """Render ALL pages of a scanned PDF and extract material via Tesseract OCR.
    Tries 'eng' first (better for factory drawings with English text),
    then falls back to 'eng+chi_sim' for Chinese-character drawings.
    Applies quality filtering to reject garbage OCR results."""
    try:
        import fitz as pymupdf_lib
        doc = pymupdf_lib.open(str(filepath))
        # Render and OCR all pages
        pix = doc[0].get_pixmap(dpi=200)
        img = Image.frombytes('RGB', [pix.width, pix.height], pix.samples)
        # Sharpen to improve OCR accuracy on factory drawings
        img = img.filter(ImageFilter.SHARPEN)
        img = ImageEnhance.Sharpness(img).enhance(1.5)

        # Try English first (better for PP6331, MOULDEX tables, etc.)
        ocr_eng = pytesseract.image_to_string(img, lang='eng')
        doc.close()
        mat_name, mat_code = _ocr_text_extract(ocr_eng)
        filtered = _quality_filter(mat_name)
        if filtered:
            return filtered, mat_code
        # Fallback: Chinese+English for drawings with Chinese text
        doc = pymupdf_lib.open(str(filepath))
        pix = doc[0].get_pixmap(dpi=200)
        img = Image.frombytes('RGB', [pix.width, pix.height], pix.samples)
        img = img.filter(ImageFilter.SHARPEN)
        img = ImageEnhance.Sharpness(img).enhance(1.5)
        ocr_chi = pytesseract.image_to_string(img, lang='eng+chi_sim')
        doc.close()
        mat_name, mat_code = _ocr_text_extract(ocr_chi)
        filtered = _quality_filter(mat_name)
        if filtered:
            return filtered, mat_code
        return '', ''
    except Exception:
        return '', ''


def _ocr_text_extract(ocr_text):
    """Extract material from OCR text — handles both clean and noisy OCR."""
    if not ocr_text or not ocr_text.strip():
        return '', ''
    # Strategy 1: Standard extraction pipeline (works on cleaner OCR)
    mat_name, mat_code = extract_material_from_text(ocr_text)
    if mat_name and _looks_like_material_value(mat_name) and '|' not in mat_name:
        return mat_name, mat_code
    # Strategy 2: Tolerant keyword scan
    mat_name = _extract_material_from_ocr_text(ocr_text)
    if mat_name and '|' not in mat_name:
        return mat_name, ''
    # Strategy 3: Resin grade code in table context (e.g. "PP6331 4.03g | MOULDEX")
    # No ^/$ anchors — grade code appears mid-line in table rows
    GRADE_PAT = re.compile(r'\b[A-Z]{2,4}-?\d{3,}[A-Z0-9-]*\b', re.IGNORECASE)
    for line in ocr_text.split('\n'):
        lu = line.strip().upper()
        if not lu or len(lu) > 150:
            continue
        # Skip noise headers
        if any(s in lu for s in ['PROPRIETARY', 'DOCUMENT', 'REVISION', 'SPECIFICATION',
                                   'DRAWING IS', 'SHOULD NOT BE', 'WITHOUT THE',
                                   'PRIOR WRITTEN', 'MANUFACTURER', 'EFFECTIVE DATE',
                                   'INCHES', 'BREAK ALL', 'SHARP EDGES', 'TOLERANCES']):
            continue
        # Look for grade codes near MOULDEX/TAITALAC/production info
        for gm in GRADE_PAT.finditer(lu):
            ctx_start = max(0, gm.start() - 40)
            ctx_end = min(len(line), gm.end() + 80)
            ctx = line[ctx_start:ctx_end].upper()
            if any(s in ctx for s in ['MOULDEX', 'TAITALAC', 'TOYOLAC', 'TERLUX',
                                        'PAGE', 'REV', 'SCALE', 'WEIGHT', 'GRADE']):
                grade = gm.group().strip()
                if len(grade) >= 3:
                    # Verify it's not a random fragment
                    kw_pat = re.compile(
                        r'(?:POLYPROPYLENE|POLYCARBONATE|ABS|PC\b|PP\b|PE\b|PVC|STEEL|'
                        r'SUS\d|BRASS|COPPER|RUBBER|SILICONE|EPDM|NYLON|PTFE|POM|TPE|TPU)',
                        re.IGNORECASE
                    )
                    surrounding = line[max(0,gm.start()-80):gm.end()+30]
                    if kw_pat.search(surrounding) or grade.startswith(('PP', 'PC', 'PE', 'ABS', 'PVC')):
                        return grade, ''
    return '', ''


def _extract_material_from_ocr_text(ocr_text):
    """Tolerant material extraction for OCR text — finds material values even when line has noise."""
    # Build alternation pattern string (joined by |)
    _pat = (r'(?<!\w)'
            r'(?:'
            r'VALOX\s+(?:PBT|ASA)|'
            r'BAKELITE\s+[A-Z0-9-]+|'
            r'LEXAN\s+[A-Z0-9-]+|'
            r'MAKROLON\s+[A-Z0-9-]+|'
            r'TAITALAC\s+[A-Z0-9-]+|'
            r'TOYOLAC\s+[A-Z0-9-]+|'
            r'TERLUX\s+[A-Z0-9]+|'
            r'LUSTRAN\s+[A-Z0-9-]+|'
            r'POLYPROPYLENE\b|'
            r'POLYCARBONATE\b|'
            r'POLYETHYLENE\b|'
            r'HIGH\s+DENSITY\s+POLYETHYLENE|'
            r'LOW\s+DENSITY\s+POLYETHYLENE|'
            r'POLYVINYL\s+CHLORIDE\b|'
            r'ACRYLONITRILE\s+BUTADIENE|'
            r'POLYISOPRENE\b|'
            r'NYLON\s+\d|'
            r'\bPOM\b|'
            r'\bPTFE\b|'
            r'\bTPE\b|'
            r'\bTPU\b|'
            r'\bEPDM\b|'
            r'\bNBR\b|'
            r'\bSILICONE\b|'
            r'\bRUBBER\b|'
            r'\bBRASS\b|'
            r'\bCOPPER\b|'
            r'\bSTEEL\b|'
            r'SUS\d+%?|'
            r'\bABS\b|'
            r'\bPC\b|'
            r'\bPP\b|'
            r'\bPE\b|'
            r'\bPVC\b|'
            r'\bPS\b|'
            r'\bPMMA\b|'
            r'\bPU\b|'
            r'\bHDPE\b|'
            r'\bLDPE\b|'
            r'\bETFE\b|'
            r'\bPVDF\b|'
            r'\bPET\b|'
            r'PC/ABS\b|'
            r'CO-POLYESTER\b|'
            r'\bPP\b|'
            r'POLYURETHANE\b|'
            r'POLYAMIDE\b|'
            r'POLYOLEFIN\b|'
            r'POLYSULFONE\b|'
            r'\bPEEK\b|'
            r'\bULTEM\b|'
            r'CYCLO-OLEFIN'
            r')')
    oc_mat = re.compile(_pat, re.IGNORECASE)
    NOISE_ENDERS = re.compile(
        r'\s+(?:CERTIFICATION|TRACEABILITY|REQUIRE|CONFIDENTIAL|PROPRIETARY|'
        r'CONDITIONALLY|DISCLOSE|REPRODUCE|ACCEPTANCE|INSPECTION|PERFORM|'
        r'DRAWING|DOCUMENT|SPECIFICATION|VISUAL|INCOMING|ENCAPSULATIONS|'
        r'APPLICATION|DIMENSION|MEASURE|CHECK|REVISION|SCALE|SHEET|SIZE|'
        r'THIRD ANGLE|UNLESS OTHERWISE|OR EQUIVALENT|SUBJECT MATTER|'
        r'WITHOUT THE|PRIOR WRITTEN|ICU MEDICAL|MANUFACTURER|EFFECTIVE DATE)\b',
        re.IGNORECASE
    )

    for line in ocr_text.split('\n'):
        lu = line.strip().upper()
        # Skip obvious non-material lines
        if any(skip in lu for skip in ['CERTIFICATION', 'TRACEABILITY', 'REQUIRE',
                                        'CONFIDENTIAL', 'PROPRIETARY', 'CONDITIONALLY',
                                        'DISCLOSE', 'REPRODUCE', 'ACCEPTANCE']):
            continue
        if 'MATERIAL CERT' in lu or 'MATERIALS:' in lu:
            continue
        # Skip "MATERIAL:" header lines with no real material value after it
        m_colon = re.match(r'MATERIAL\s*:\s*(.+)$', lu)
        if m_colon:
            after = m_colon.group(1).strip()
            if not after or any(skip in after for skip in
                ['VISUAL', 'PHYSICAL', 'CERTIFICATION', 'TRACEABILITY', 'PERFORM',
                 'ACCEPTANCE', 'SPECIFIED', 'CONDITIONS', 'CRITERIA']):
                continue
        # Find the material match and take trailing portion
        m = oc_mat.search(lu)
        if m:
            idx = m.start()
            # Grab up to 200 chars starting from the keyword (generous for long specs)
            candidate = line[idx:idx+200].strip()
            # Truncate at noise boundary
            noise_match = NOISE_ENDERS.search(candidate)
            if noise_match:
                candidate = candidate[:noise_match.start()].strip()
            # Strip trailing punctuation/garbage
            candidate = re.sub(r'\s*[|/\\:,;#].*$', '', candidate).strip()
            if len(candidate) >= 3:
                return candidate
    return ''


def extract_material_from_ocr(filepath, existing_text=''):
    """Try OCR fallback when text extraction yields insufficient content.
    Returns (mat_name, mat_code, method_label)."""
    mat_name, mat_code = _ocr_extract_material(filepath)
    if mat_name:
        return mat_name, mat_code, 'ocr'
    # Also check if text was empty → truly scanned
    return '', '', 'scanned_no_text'

def extract_description_from_text(text, part_no=''):
    """Extract full part description from PDF text using multi-pattern matching."""
    lines = [l.strip() for l in text.split('\n') if l.strip()]

    # Strategy 1: DESCRIPTION & PART # header followed by description
    for j, line in enumerate(lines):
        if re.match(r'DESCRIPTION\s*&\s*PART\s*#', line, re.IGNORECASE):
            for k in range(j+1, min(j+4, len(lines))):
                d = lines[k].strip()
                if d and len(d) > 5 and not any(n in d.upper() for n in DESC_NOISE):
                    return _clean_description(d)
            break

    # Strategy 2: Look in the last third of text (title block is usually at bottom)
    n = len(lines)
    search_area = lines[max(0, n//3):]
    for line in reversed(search_area):
        # Pattern: full description ending with PN, e.g. "BAG SPIKE, ..., R1-15460"
        if DESC_PN_PATTERN.search(line) and len(line) > 15:
            # Check it has descriptive keywords, not just dimensions
            has_keyword = any(kw in line.upper() for kw in DESC_KEYWORDS)
            has_dim = any(kw in line.upper() for kw in ['±', 'MM', 'INCH', 'TOLERANCE', 'DECIMAL'])
            if has_keyword and not has_dim:
                # Strip trailing PN
                desc = re.sub(r',?\s*(?:R1-\d+|[A-Z]\d+-\d+)\s*$', '', line).strip()
                if desc and len(desc) > 5:
                    return _clean_description(desc)

    # Strategy 3: "DESCRIPTIVE TEXT (PN)" pattern
    for line in lines:
        m = re.match(r'^([A-Z][A-Z ,.\'\-/]+)\s*\((R1-\d+|[A-Z]\d+-\d+)\)\s*$', line)
        if m and len(m.group(1)) > 5:
            return _clean_description(m.group(1).strip())

    # Strategy 4: "PN DESCRIPTIVE TEXT" pattern (PN first, then description)
    for line in lines:
        m = re.match(r'^((?:R1-\d+|[A-Z]\d+-\d+))\s+(.+)$', line)
        if m:
            rest = m.group(2).strip()
            if any(kw in rest.upper() for kw in DESC_KEYWORDS) and 5 < len(rest) < 100:
                if not any(n in rest.upper() for n in DESC_NOISE):
                    return _clean_description(rest)

    # Strategy 5: Standalone description line near part number mention
    for j, line in enumerate(lines):
        if DESC_PN_PATTERN.search(line):
            # Check surrounding lines
            for k in range(max(0, j-2), min(len(lines), j+3)):
                candidate = lines[k].strip()
                if (candidate and len(candidate) > 8 and len(candidate) < 150
                        and any(kw in candidate.upper() for kw in DESC_KEYWORDS)
                        and not any(n in candidate.upper() for n in DESC_NOISE)):
                    return _clean_description(candidate)

    return ''

def _clean_description(desc):
    """Clean up a raw description string."""
    # Remove trailing comma
    desc = desc.rstrip(',').strip()
    # Remove duplicate PN at end
    desc = re.sub(r',?\s*(?:R1-\d+|[A-Z]\d+-\d+)\s*$', '', desc).strip()
    return desc

# ── Fast-path extraction (no LLM) ─────────────────────────────────────────
def fast_extract(filepath, icu_map):
    """Quick regex-based extraction with text-based material fallback."""
    filename = filepath.name
    stem = filepath.stem.replace('_mdx', '')
    folder = get_folder_type(filepath)

    result = {
        'fileName': filename,
        'filePath': str(filepath),
        'drawingNo': '', 'revision': '', 'partNo': '',
        'description': '', 'color': '',
        'materialName': '', 'materialCode': '',
        'category': '零件', 'bom': [],
        'source': folder,
        'method': 'fast_regex',
    }

    # Filename-based extraction — broader patterns
    fn_patterns = [
        # Part numbers with hyphen: R1-16529, B-077, E09-000-642
        (r'([A-Z]\d{0,2}-\d{2,5}(-\d{2,4}){0,3})', 'partNo'),
        # Part numbers without hyphen: CP95004
        (r'([A-Z]{2,3}\d{5,})', 'partNo'),
        (r'(?:^|[_\-])([A-Z]{2,4}-\d{4,})(?:_|\))', 'partNo'),
        (r'^(\d{2,5}-\d+)', 'partNo'),
        # Rev patterns — applied to both filename and stem
        (r'Rev\.?\.?([A-Z]?\.?\d*)', 'revision'),
        (r'\((Rev\.?[A-Z]\d*)\)', 'revision'),
        # ICU style: XX-XXXX-MC_08 or XX-XXXX-MC_08_mdx -> revision 08
        (r'[-_](\d{2})\.?(?:\.pdf|$|_mdx)', 'revision'),
    ]
    for pat, field in fn_patterns:
        # Try on filename first, then on stem (without extension)
        for text in [filename, stem]:
            m = re.search(pat, text)
            if m and not result[field]:
                val = m.group(1)
                if not re.match(r'^(Rev\.?|rev\.?)', val, re.IGNORECASE):
                    result[field] = val
                    break

    # ICU lookup for additional fields — broader key matching
    icu_matched = False
    if folder.startswith('icu') and not result['materialName']:
        # Build candidate keys: exact stem, underscore-split, and regex-extracted base PN
        candidates = set()
        candidates.add(stem)
        candidates.add(stem.split('_')[0])
        # Extract base PN from patterns like "27-0246-MC_08" or "R1-8392-MC-01"
        m = re.match(r'^([A-Z]?\d+-\d+)', stem.replace('_', '-'))
        if m:
            candidates.add(m.group(1))
        # Also try just the numeric prefix for complex stems like "C74-49554-MC_05_C74-49554"
        m2 = re.match(r'^([A-Z]\d+-\d+(?:-\d+)?)', stem.replace('_', '-'))
        if m2:
            candidates.add(m2.group(1))
        for key in candidates:
            if key in icu_map:
                icu_item = icu_map[key]
                raw_mat = icu_item.get('material', '')
                # Extract material code (R-prefix or number pattern) from material string
                code_match = re.search(r'COMMODITY\s*(?:NO\.|#)?\s*([A-Z]?\d+-\d+)', raw_mat, re.IGNORECASE)
                result['materialCode'] = code_match.group(1) if code_match else ''
                result['materialName'] = raw_mat
                result['description'] = icu_item.get('nameEN', '')
                result['color'] = icu_item.get('color', '')
                # Preserve revision from filename if ICU has none
                if not result['revision']:
                    result['revision'] = ''  # ICU lookup doesn't provide revision
                result['method'] = 'fast_regex+icu_lookup'
                icu_matched = True
                break

        # Text-based extraction (material + description) — runs for ALL files
        try:
            import fitz
            doc = fitz.open(str(filepath))
            text = ''.join(page.get_text() for page in doc)
            doc.close()

            # Material — only if not already found via ICU
            if not result['materialName']:
                mat_name, mat_code = extract_material_from_text(text)
                if mat_name:
                    result['materialName'] = mat_name
                    if mat_code and not result['materialCode']:
                        result['materialCode'] = mat_code

            # Description — only if still empty or suspiciously short (ICU truncation)
            if not result['description'] or len(result['description']) < 10:
                txt_desc = extract_description_from_text(text, result['partNo'])
                if txt_desc and len(txt_desc) >= 5:
                    result['description'] = txt_desc

            # DrawingNo — from ICU dwgNo or text patterns
            if not result['drawingNo'] and folder.startswith('icu'):
                # Try ICU dwgNo
                base_pn = result['partNo']
                if base_pn and base_pn in icu_map:
                    dwg = icu_map[base_pn].get('dwgNo', '')
                    if dwg:
                        result['drawingNo'] = dwg
                # Try text pattern: DRAWING #: XXX or DOCUMENT #: XXX
                if not result['drawingNo']:
                    for line in text.split('\n'):
                        m = re.search(r'(?:DRAWING|DOCUMENT)\s*#?\s*[:：]?\s*(R1-\d+|[A-Z]\d+-\d{2,}|[A-Z]{2,4}-\d{2,})', line, re.IGNORECASE)
                        if m:
                            result['drawingNo'] = m.group(1)
                            break

            if not icu_matched:
                result['method'] = 'fast_regex+text_extract'
            elif not result['description'] or len(result['description']) < 10:
                # ICU provided material but description was incomplete
                result['method'] = 'fast_regex+icu_lookup+text_desc'
        except Exception:
            pass

    return result

# ── LLM extraction ─────────────────────────────────────────────────────────
def call_llm(prompt):
    """Call agnes CLI and return parsed result or None."""
    try:
        result = subprocess.run(
            ['npx', 'agnes', 'text', 'chat', '--model', 'agnes-2.0-flash', '--json', '--prompt', prompt],
            capture_output=True, text=True, timeout=120,
            encoding='utf-8', errors='replace', shell=True
        )
        if result.returncode != 0:
            return None

        # Parse JSON response
        try:
            resp = json.loads(result.stdout)
        except json.JSONDecodeError:
            return None

        if isinstance(resp, dict) and 'text' in resp:
            return parse_llm_text(resp['text'])
        return resp if isinstance(resp, dict) else None
    except subprocess.TimeoutExpired:
        return None
    except Exception as e:
        print(f"    LLM call error: {e}")
        return None

def llm_extract_drawing(text, filename, icu_map, folder):
    """Use LLM to extract structured data from drawing text."""
    prompt = LLM_PROMPT_TEMPLATE.format(text=text)
    resp = call_llm(prompt)
    if not resp or not isinstance(resp, dict):
        return None

    partNo = resp.get('partNo', '').strip()
    drawingNo = resp.get('drawingNo', '').strip()
    description = resp.get('description', '').strip()

    # Skip if we got garbage (common failure mode)
    noise_words = ['REVISION:', '2:1', 'SHEET:', 'DIMENSIONS', 'CHECK', 'TOLERANCE',
                   'SCALE', 'SURFACE TREATMENT', 'PROCESS', 'STANDARD', 'NOTE',
                   'NON-LOCKING', 'ENCAPSULATIONS', 'UNIVERSAL', 'DRAWING', 'ASSEMBLY',
                   'REVISION HISTORY', 'QUALITY', 'SPECIFICATION', 'PERFORM',
                   'INCOMING INSPECTION', 'ACCEPTANCE CRITERIA']
    if description:
        desc_upper = description.upper()
        if any(desc_upper.startswith(nw) or desc_upper == nw for nw in noise_words):
            return None

    result = {
        'fileName': filename,
        'filePath': '',
        'drawingNo': drawingNo,
        'revision': resp.get('revision', '').strip(),
        'partNo': partNo,
        'description': description,
        'color': resp.get('color', '').strip(),
        'materialName': resp.get('materialName', '').strip(),
        'materialCode': resp.get('materialCode', '').strip(),
        'category': '零件',
        'source': folder,
        'bom': [],
        'method': 'llm_extract',
    }
    return result

def llm_extract_bom(text):
    """Use LLM to extract BOM from drawing text."""
    prompt = BOM_PROMPT_TEMPLATE.format(text=text)
    resp = call_llm(prompt)
    if not resp:
        return []
    if isinstance(resp, dict):
        resp = [resp]
    if not isinstance(resp, list):
        return []

    bom = []
    seen = set()
    for item in resp:
        if not isinstance(item, dict):
            continue
        pn = item.get('partNo', '').strip()
        if not pn or len(pn) < 3:
            continue
        if pn in seen:
            continue
        seen.add(pn)
        bom.append({
            'partNo': pn,
            'description': item.get('description', '').strip(),
            'material': item.get('material', '').strip(),
            'qty': str(item.get('qty', '')).strip(),
        })
    return bom

# ── Regex BOM extraction (fast path for assembly files) ─────────────────────
def extract_bom_from_text(text):
    """Extract BOM entries from PDF text using regex patterns."""
    bom = []
    seen = set()

    # Pattern 1: 'NAME (PN)' — most common ICU format
    for m in re.finditer(r'([A-Z][A-Z ,.\'\-/]+?)\s*\((R1-\d+|[A-Z]\d+-\d+|\d{2,5}-\d+)\)', text):
        desc = m.group(1).strip()
        pn = m.group(2).strip()
        if pn not in seen and len(desc) > 2 and len(desc) < 80:
            seen.add(pn)
            bom.append({'partNo': pn, 'description': desc, 'material': '', 'qty': ''})

    # Pattern 2: NAME on one line, PN on next (multi-line BOM)
    lines = text.split('\n')
    for j in range(len(lines) - 1):
        l1 = lines[j].strip()
        l2 = lines[j + 1].strip()
        if (re.match(r'^[A-Z][A-Z ,.\'/\-]+$', l1)
                and re.search(r'(?:^|[\s(])R1-\d+', l2)):
            pn_m = re.search(r'(R1-\d+)', l2)
            if pn_m:
                pn = pn_m.group(1)
                if pn not in seen and len(l1) > 2:
                    seen.add(pn)
                    bom.append({'partNo': pn, 'description': l1, 'material': '', 'qty': ''})

    # Pattern 3: 'PART #: ...' followed by description lines
    for j, line in enumerate(lines):
        if re.match(r'PART\s*#?\s*:', line.strip(), re.IGNORECASE):
            for k in range(j + 1, min(j + 5, len(lines))):
                candidate = lines[k].strip()
                pn_m = re.search(r'(R1-\d+|[A-Z]\d+-\d+)', candidate)
                if pn_m:
                    desc = re.sub(r'\s*\(R1-\d+\)\s*', ' ', candidate).strip()
                    if desc and desc not in seen:
                        seen.add(pn_m.group(1))
                        bom.append({'partNo': pn_m.group(1), 'description': desc, 'material': '', 'qty': ''})

    return bom

# ── Main pipeline ───────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("PN-Lookup PDF Drawing Extractor v7 (LLM Semantic)")
    print("=" * 60)

    # Load ICU lookup
    icu_map = {}
    if ICU_LOOKUP.exists():
        try:
            with open(ICU_LOOKUP, encoding='utf-8') as f:
                for item in json.load(f):
                    pn = item.get('partNo', '')
                    if pn:
                        icu_map[pn] = item
        except Exception:
            pass
    print(f"ICU lookup: {len(icu_map)} entries")

    # Get all PDF files
    pdf_files = sorted(glob.glob(str(BASE) + '/**/*.pdf', recursive=True))
    print(f"Found {len(pdf_files)} PDF files")

    results = []
    llm_count = 0
    fast_count = 0
    scanned_count = 0
    error_count = 0
    start_time = time.time()
    failed_texts = {}  # Store texts for retry

    for i, fp in enumerate(pdf_files):
        if (i + 1) % 50 == 0:
            elapsed = time.time() - start_time
            rate = (i + 1) / max(elapsed, 1)
            eta = (len(pdf_files) - i - 1) / max(rate, 0.001) / 60
            print(f"  Progress: {i+1}/{len(pdf_files)} ({eta:.1f}min ETA)")

        filename = Path(fp).name
        folder = get_folder_type(fp)

        # Fast path: try filename regex
        fast = fast_extract(Path(fp), icu_map)
        has_pn = bool(fast.get('partNo'))
        has_dwg = bool(fast.get('drawingNo'))

        # Text-based enhancement (material + description + drawingNo) — ALWAYS runs
        try:
            import fitz
            doc = fitz.open(str(fp))
            text = ''.join(page.get_text() for page in doc)
            doc.close()

            if not fast['materialName']:
                mat_name, mat_code = extract_material_from_text(text)
                if mat_name:
                    fast['materialName'] = mat_name
                    if mat_code and not fast['materialCode']:
                        fast['materialCode'] = mat_code
                # OCR fallback for scanned PDFs or files with text but no material match
                if not fast['materialName']:
                    ocr_name, ocr_code, ocr_method = extract_material_from_ocr(
                        fp, existing_text=text)
                    if ocr_name:
                        fast['materialName'] = ocr_name
                        if ocr_code and not fast['materialCode']:
                            fast['materialCode'] = ocr_code
                        fast['_ocr_used'] = True

            if not fast['description'] or len(fast['description']) < 10:
                txt_desc = extract_description_from_text(text, fast['partNo'])
                if txt_desc and len(txt_desc) >= 5:
                    fast['description'] = txt_desc

            if not fast['drawingNo'] and folder.startswith('icu'):
                base_pn = fast['partNo']
                if base_pn and base_pn in icu_map:
                    dwg = icu_map[base_pn].get('dwgNo', '')
                    if dwg:
                        fast['drawingNo'] = dwg
                if not fast['drawingNo']:
                    for line in text.split('\n'):
                        m = re.search(r'(?:DRAWING|DOCUMENT)\s*#?\s*[:：]?\s*(R1-\d+|[A-Z]\d+-\d{2,}|[A-Z]{2,4}-\d{2,})', line, re.IGNORECASE)
                        if m:
                            fast['drawingNo'] = m.group(1)
                            break

            # Regex BOM for ICU assembly files
            if folder.startswith('icu') and not fast.get('bom'):
                reg_bom = extract_bom_from_text(text)
                if reg_bom:
                    fast['bom'] = reg_bom

            # Update method label
            method_parts = ['fast_regex']
            if 'icu_lookup' in fast.get('method', ''):
                method_parts.append('icu_lookup')
            if fast['_ocr_used']:
                method_parts.append('ocr')
            elif fast['materialName'] or fast['description'] or fast['drawingNo']:
                method_parts.append('text_extract')
            if fast.get('bom'):
                method_parts.append('regex_bom')
            fast['method'] = '+'.join(method_parts)
        except Exception:
            pass

        if has_pn or has_dwg:
            results.append(fast)
            fast_count += 1
            continue

        # Slow path: needs LLM — extract text
        text = ''
        try:
            import fitz
            doc = fitz.open(str(fp))
            text = ''.join(page.get_text() for page in doc)
            doc.close()
        except Exception:
            pass

        if not text:
            # Try OCR before giving up
            ocr_name, ocr_code, ocr_method = extract_material_from_ocr(fp, existing_text=text)
            if ocr_name:
                result = {
                    'fileName': filename, 'filePath': str(fp),
                    'drawingNo': '', 'revision': '', 'partNo': '',
                    'description': '', 'color': '',
                    'materialName': ocr_name, 'materialCode': ocr_code,
                    'category': '零件', 'bom': [],
                    'source': folder,
                    'method': 'ocr',
                }
            else:
                # Truly scanned/no-text file
                result = {
                    'fileName': filename, 'filePath': str(fp),
                    'drawingNo': '', 'revision': '', 'partNo': '',
                    'description': '', 'color': '',
                    'materialName': '', 'materialCode': '',
                    'category': '零件', 'bom': [],
                    'source': folder,
                    'method': 'scanned_no_text',
                }
            results.append(result)
            if ocr_method == 'scanned_no_text':
                scanned_count += 1
            continue

        # Store text for potential retry
        failed_texts[filename] = text

        # Try LLM extraction
        llm_result = None
        try:
            llm_result = llm_extract_drawing(text, filename, icu_map, folder)
        except Exception as e:
            print(f"    LLM error for {filename}: {e}")

        if llm_result and (llm_result.get('partNo') or llm_result.get('description')):
            llm_result['filePath'] = str(fp)
            llm_result['bom'] = []

            # Also try BOM extraction
            try:
                bom = llm_extract_bom(text)
                if bom:
                    llm_result['bom'] = bom
                    llm_result['method'] = 'llm_extract+bom'
            except Exception:
                pass

            results.append(llm_result)
            llm_count += 1
            continue

        # Fallback: fast regex only (might have partial data from filename)
        fast['filePath'] = str(fp)
        fast['bom'] = []
        results.append(fast)
        error_count += 1

    elapsed = time.time() - start_time

    # ── Partner fallback: copy material from same-base files ───────────────
    def _norm_fn(fn):
        n = fn.lower()
        n = re.sub(r'\(rev\.[^)]+\)', '', n)
        n = re.sub(r'-c\.pdf$', '.pdf', n)
        n = n.replace('_mdx.pdf', '.pdf')
        n = re.sub(r'_(\d{1,2})\.pdf$', '.pdf', n)
        return n

    # Build valid-material lookup
    _valid_mats = {}
    for r in results:
        if r.get('materialName') and _looks_like_material_value(r['materialName']) and _has_material_keyword(r['materialName']) and _material_starts_valid(r['materialName']):
            _valid_mats[_norm_fn(r['fileName'])] = (r['materialName'], r.get('materialCode', ''))

    from collections import defaultdict
    _by_norm = defaultdict(list)
    for r in results:
        _by_norm[_norm_fn(r['fileName'])].append(r)

    partner_copied = 0
    for norm_key, members in _by_norm.items():
        has_mat = [m for m in members if m.get('materialName') and _norm_fn(m['fileName']) == norm_key]
        lacks_mat = [m for m in members if not m.get('materialName')]
        for lacking in lacks_mat:
            if has_mat and norm_key in _valid_mats:
                mat_name, mat_code = _valid_mats[norm_key]
                lacking['materialName'] = mat_name
                if not lacking.get('materialCode'):
                    lacking['materialCode'] = mat_code
                partner_copied += 1

    print(f"\nDone in {elapsed/60:.1f} min")
    print(f"  Fast path: {fast_count}")
    print(f"  LLM path:  {llm_count}")
    print(f"  Scanned:   {scanned_count}")
    print(f"  Errors:    {error_count}")
    if partner_copied:
        print(f"  Partner copy: {partner_copied}")

    # Stats
    has_pn = sum(1 for r in results if r['partNo'])
    has_dwg = sum(1 for r in results if r['drawingNo'])
    has_mat = sum(1 for r in results if r['materialName'])
    has_desc = sum(1 for r in results if r['description'])
    has_rev = sum(1 for r in results if r['revision'])
    has_bom = sum(1 for r in results if r.get('bom'))

    print(f"\n=== Stats ===")
    print(f"  partNo:     {has_pn}/{len(results)} ({has_pn/len(results)*100:.1f}%)")
    print(f"  drawingNo:  {has_dwg}/{len(results)} ({has_dwg/len(results)*100:.1f}%)")
    print(f"  material:   {has_mat}/{len(results)} ({has_mat/len(results)*100:.1f}%)")
    print(f"  description:{has_desc}/{len(results)} ({has_desc/len(results)*100:.1f}%)")
    print(f"  revision:   {has_rev}/{len(results)} ({has_rev/len(results)*100:.1f}%)")
    print(f"  BOM files:  {has_bom}")

    methods = {}
    for r in results:
        m = r.get('method', 'unknown')
        methods[m] = methods.get(m, 0) + 1
    print(f"\nMethods:")
    for k, v in sorted(methods.items(), key=lambda x: -x[1]):
        print(f"  {v:4d}  {k}")

    # Save JSON
    output = {
        'generatedAt': time.strftime('%Y-%m-%dT%H:%M:%S'),
        'totalFiles': len(results),
        'items': results,
    }
    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"\nJSON saved: {OUT_JSON}")

    # Save scanned list
    scanned = [r['fileName'] for r in results if r['method'] == 'scanned_no_text']
    with open(OUT_SCAN, 'w', encoding='utf-8') as f:
        json.dump(scanned, f, ensure_ascii=False, indent=2)
    print(f"Scanned list: {len(scanned)} files")

    # Save failed texts for potential retry
    if failed_texts:
        retry_path = OUT_JSON.parent / 'failed_texts_v7.json'
        with open(retry_path, 'w', encoding='utf-8') as f:
            json.dump(failed_texts, f, ensure_ascii=False, indent=2)
        print(f"Failed texts saved: {len(failed_texts)} files -> {retry_path}")

    # Save Excel
    write_excel(results, OUT_XLSX)

def write_excel(results, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '圖面資料'
    headers = ['圖檔名', '圖號(Drawing No.)', '版本', '品號(Part No.)', '品名(Description)',
               '顏色(Color)', '原料名稱(Material Name)', '原料編碼(Material Code)',
               '分類(Category)', '提取方法(Method)', '來源資料夾(Folder)', '檔案路徑(FilePath)']
    ws.append(headers)
    for h in ws[1]:
        h.font = Font(bold=True, color='FFFFFF')
        h.fill = PatternFill('solid', fgColor='4472C4')
        h.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for r in results:
        ws.append([
            r['fileName'],
            r.get('drawingNo', ''),
            r.get('revision', ''),
            r.get('partNo', ''),
            (r.get('description', '') or '')[:100],
            r.get('color', '') or '',
            (r.get('materialName', '') or '')[:120],
            r.get('materialCode', '') or '',
            r.get('category', '零件'),
            r.get('method', ''),
            r.get('source', ''),
            r.get('filePath', ''),
        ])

    col_widths = [40, 22, 8, 20, 50, 12, 50, 20, 12, 14, 25, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Red highlight for scanned_no_text
    scanned_set = set(r['fileName'] for r in results if r.get('method') == 'scanned_no_text')
    red_fill = PatternFill('solid', fgColor='FFE0E0')
    for row in ws.iter_rows(min_row=2):
        if row[0].value in scanned_set:
            for cell in row:
                cell.fill = red_fill

    # BOM sheet
    ws2 = wb.create_sheet('組件BOM')
    ws2.append(['組件圖檔名', '組件品號', '組件分類', '組成零件品號', '組成零件品名', '組成零件原料', '單位用量'])
    for h in ws2[1]:
        h.font = Font(bold=True, color='FFFFFF')
        h.fill = PatternFill('solid', fgColor='4472C4')
        h.alignment = Alignment(horizontal='center', vertical='center')
    for r in results:
        if r.get('bom'):
            for b in r['bom']:
                ws2.append([r['fileName'], r.get('partNo', '') or r.get('drawingNo', ''),
                            r.get('category', '零件'),
                            b.get('partNo', ''), b.get('description', ''),
                            b.get('material', ''), b.get('qty', '')])
    for i, w in enumerate([35, 20, 10, 20, 40, 50, 8], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Scanned sheet
    ws3 = wb.create_sheet('掃描圖檔(需手動處理)')
    ws3.append(['圖檔名', 'OCR結果', '來源資料夾'])
    for h in ws3[1]:
        h.font = Font(bold=True, color='FFFFFF')
        h.fill = PatternFill('solid', fgColor='C00000')
    for r in results:
        if r.get('method') == 'scanned_no_text':
            ws3.append([r['fileName'], '(純圖片無文字層)', r.get('source', '')])
    for i, w in enumerate([40, 25, 20], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    print(f"Excel saved: {path}")

if __name__ == '__main__':
    main()
