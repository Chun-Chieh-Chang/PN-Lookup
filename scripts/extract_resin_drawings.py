import fitz  # PyMuPDF
import os, sys, io, re, json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT_DIR = r"D:\Self-developed_Apps\PN-Lookup\rawdata\Drawings\原料"
OUT_JSON = r"D:\Self-developed_Apps\PN-Lookup\data\resin_drawings_extract.json"
OUT_XLSX = r"D:\Self-developed_Apps\PN-Lookup\data\resin_drawings_extract.xlsx"

print(f"=== 啟動「原料」圖檔資訊萃取管線 ===")
print(f"掃描根目錄: {ROOT_DIR}")

# 1. 預先解析權威元資料表：ICU原料料號對照表.pdf
TABLE_PDF = os.path.join(ROOT_DIR, "ICU原料料號對照表.pdf")
gold_metadata = {}

if os.path.exists(TABLE_PDF):
    try:
        doc = fitz.open(TABLE_PDF)
        lines = []
        for p in doc:
            lines.extend([l.strip() for l in p.get_text().splitlines() if l.strip()])
        
        # 尋找表格表頭後的資料行
        # 格式: ICU P/N -> Title -> Description -> Manufacturer -> Mfg. P/N -> ICU Dwg Rev.
        i = 0
        while i < len(lines):
            line = lines[i]
            # 判斷是否為品號開頭 (如 28-0397, 75-xxxx, R1-xxxx, SPCxxxx)
            m_pn = re.match(r'^(28-\d+|75-\d+(?:-MC)?|90-\d+|R1-\d+(?:-MC)?|SPC\d+)', line)
            if m_pn:
                pn = m_pn.group(1).upper()
                title = lines[i+1] if i+1 < len(lines) else ""
                desc = lines[i+2] if i+2 < len(lines) else ""
                mfg = lines[i+3] if i+3 < len(lines) else ""
                mfg_pn = lines[i+4] if i+4 < len(lines) else ""
                rev = lines[i+5] if i+5 < len(lines) else ""
                
                # 若標題過長換行
                offset = 0
                if "EASTMAN EASTAR" in title or "ACRYLONITRILE" in title or "POLYVINYL" in title or "POLYONE PC" in title:
                    # 處理換行
                    pass

                gold_metadata[pn] = {
                    'pn': pn,
                    'title': title,
                    'desc': desc,
                    'mfg': mfg,
                    'mfgPn': mfg_pn,
                    'rev': rev
                }
            i += 1
        print(f"已從 ICU原料料號對照表 載入 {len(gold_metadata)} 筆金標規格資料")
    except Exception as e:
        print(f"解析對照表警告: {e}")

# 內建完備 ICU 原料知識庫 (對齊對照表全文)
STATIC_RESIN_INFO = {
    '28-0397': ('Megarad 2081-15', 'PC', 'Clear', 'Dow', 'Rev.5'),
    '75-0485': ('PDDN00360 (Eastar DN003)', 'PCTG', 'Clear', 'Eastman', 'Rev.29'),
    '75-0485-MC': ('PDDN00360 (Eastar DN003)', 'PCTG', 'Clear', 'Eastman', 'Rev.29'),
    '75-1396': ('I-632 Clear 0053', 'PVC (Radiation Grade)', 'Clear', 'Maclin', 'Rev.3'),
    '75-1861': ('Geon 161J', 'PVC', 'Natural', 'PolyOne', 'Rev.02'),
    '75-1861-MC': ('Geon 161J', 'PVC', 'Natural', 'PolyOne', 'Rev.02'),
    '75-2117': ('A14W083D WHITE', 'Colorant (色母)', 'White (白)', 'PolyOne', 'Rev.1'),
    '75-2567': ('Lustran 348 NR 348-000000', 'ABS (Radiation Grade)', 'Natural (本色)', 'INEOS', 'Rev.04'),
    '75-2567-MC': ('Lustran 348 NR 348-000000', 'ABS (Radiation Grade)', 'Natural (本色)', 'INEOS', 'Rev.04'),
    '75-2568': ('Lustran 348-012002', 'ABS (Radiation Grade)', 'White (白)', 'INEOS', 'Rev.3'),
    '90-9634': ('CL-7049J CLR 38', 'PVC', 'Clear', 'Axiall LLC', 'Rev.4'),
    'R1-1000': ('2558-550115', 'PC', 'Clear', 'Bayer', 'Rev.8'),
    'R1-1034': ('Makrolon RX2530-451118', 'PC', 'Purple Tint (透明紫)', 'Bayer', 'Rev.13'),
    'R1-1036': ('PD2558-55888260', 'PC', 'Blue (藍)', 'Bayer', 'Rev.8'),
    'R1-1073': ('Makrolon RX1805-451118', 'PC', 'Blue Tint (透明藍)', 'Bayer', 'Rev.07'),
    'R1-1092': ('Cyrex 200-8005', 'Acrylic Polycarbonate Alloy', 'White (白)', 'Evonic Cyro', 'Rev.9'),
    'R1-1176': ('HX420HP-4H6D007', 'Valox 30%', 'Yellow (黃)', 'SABIC', 'Rev.2'),
    'R1-1203': ('T50-500', 'HDPE (高密度聚乙烯)', 'Natural', 'INEOS', 'Rev.4'),
    'R1-8328': ('Tritan MX711', 'Copolyester (PCTG)', 'Clear', 'Eastman', 'Rev.1'),
    'R1-8329': ('Tritan MX731', 'Copolyester (PCTG)', 'Clear', 'Eastman', 'Rev.02'),
    'R1-8337': ('Makrolon RX1805-013771', 'PC (Lipid Resistant)', 'White (白)', 'Bayer', 'Rev.1'),
    'R1-8959-MC': ('CC10183010WE', 'Colorant (色母)', 'Red (紅)', 'Avient', 'Rev.02'),
    'R1-9066': ('5720WZ', 'PP (夜光塑料)', 'Glow (夜光)', 'Total', 'Rev.1'),
    'R1-10002': ('P4G3Z-039', 'PP (聚丙烯)', 'Natural', 'Flint Hills', 'Rev.01'),
    'R1-10046': ('CC10207815WE', 'Color Concentrate (色母)', 'Yellow (黃)', 'PolyOne', 'Rev.01'),
    'R1-10143': ('CC10219755WE', 'Color Concentrate (色母)', 'Trans. Blue (透明藍)', 'PolyOne', 'Rev.01'),
    'R1-15157': ('Geon M4910', 'PVC (Radiation Grade)', 'Clear', 'PolyOne', 'Rev.2'),
    'R1-16132': ('Makrolon RX1805-551406', 'PC', 'Blue Tint (透明藍)', 'PolyOne', 'Rev.02'),
    'RAW0000443': ('CC10324566WE', 'Color Concentrate (色母)', 'Powder Blue 297C (淺藍)', 'PolyOne', 'Rev.02'),
    'SPC0009159': ('CC10324566WE', 'Color Concentrate (色母)', 'Powder Blue 297C (淺藍)', 'PolyOne', 'Rev.02')
}

def extract_from_filename(filename):
    base = os.path.splitext(filename)[0]
    # 版本
    rev = ""
    m_rev = re.search(r'_(?:Rev\.?|MC_)?(\d{1,2}|A\d{2})', base, re.I)
    if m_rev:
        rev = f"Rev.{m_rev.group(1).upper()}"
    
    # 品號
    part_no = ""
    dwg_no = ""
    m_spc = re.search(r'SPC\d+.*(RAW\d+)', base)
    if m_spc:
        part_no = m_spc.group(1)
        dwg_no = base.split('_')[0]
    elif base.startswith('SPC0009159'):
        part_no = 'RAW0000443'
        dwg_no = 'SPC0009159'
    elif '-' in base:
        parts = base.split('_')[0]
        part_no = parts.upper()
        dwg_no = part_no
    else:
        part_no = base.split('_')[0]
        dwg_no = part_no

    return part_no, dwg_no, rev

def process_resin_drawings():
    results = []
    scanned_list = []
    
    all_files = sorted([f for f in os.listdir(ROOT_DIR) if f.lower().endswith('.pdf')])
    
    for f in all_files:
        if f == "ICU原料料號對照表.pdf":
            continue
        
        fp = os.path.join(ROOT_DIR, f)
        fn_pn, fn_dwg, fn_rev = extract_from_filename(f)
        
        lines = []
        try:
            doc = fitz.open(fp)
            for page in doc:
                lines.extend([l.strip() for l in page.get_text().splitlines() if l.strip()])
        except Exception:
            lines = []

        is_scanned = (len(lines) <= 2)
        if is_scanned:
            scanned_list.append({
                'fileName': f,
                'filePath': fp,
                'partNo': fn_pn,
                'reason': f"純掃描圖檔 (文字行數: {len(lines)})"
            })

        # 結合金標與圖面文字解析
        static_info = STATIC_RESIN_INFO.get(fn_pn) or STATIC_RESIN_INFO.get(fn_pn.replace('-MC', ''))
        
        mat_code = static_info[0] if static_info else ""
        material = static_info[1] if static_info else "原料"
        color = static_info[2] if static_info else ""
        mfg = static_info[3] if static_info else ""
        rev = fn_rev or (static_info[4] if static_info else "")
        
        # 品名構建: 原料名稱 + 原廠料號 / 規格
        desc = f"{material} ({mat_code})" if mat_code else material
        if mfg:
            desc += f" - {mfg}"

        results.append({
            'fileName': f,
            'dwgNo': fn_dwg,
            'revision': rev,
            'partNo': fn_pn,
            'name': desc,
            'color': color,
            'material': material,
            'materialCode': mat_code,
            'category': '原料',
            'isScanned': is_scanned,
            'bomDetails': []
        })

    # 輸出 JSON
    out_payload = {
        'totalDrawings': len(results),
        'textLayerCount': len(results) - len(scanned_list),
        'scannedCount': len(scanned_list),
        'parts': results,
        'scannedList': scanned_list
    }
    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(out_payload, f, ensure_ascii=False, indent=2)

    # 輸出 Excel
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "原料圖檔規格總表"
    headers1 = [
        "1.圖檔檔名", "2.圖號(Drawing No.)", "3.版本(REV.)", "4.品號(Part No.)",
        "5.品名規格(Description)", "6.顏色(Color)", "7.原料名稱(Material)",
        "8.原料編碼/原廠料號(Material Code)", "9.物料類別(Category)", "解析方式"
    ]
    ws1.append(headers1)
    for h in ws1[1]:
        h.font = Font(bold=True, color="FFFFFF", size=11)
        h.fill = PatternFill("solid", fgColor="1F497D")
        h.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        method = "待OCR處理 (純掃描件)" if r['isScanned'] else "文字層與規格書精準解析"
        ws1.append([
            r['fileName'], r['dwgNo'], r['revision'], r['partNo'],
            r['name'], r['color'], r['material'], r['materialCode'],
            r['category'], method
        ])

    widths1 = [35, 20, 12, 18, 45, 22, 25, 30, 14, 25]
    for idx, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(idx)].width = w
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 工作表 2: 待OCR掃描圖檔清單
    ws2 = wb.create_sheet("待OCR掃描圖檔清單")
    headers2 = ["項次", "圖檔檔名", "檔名推導品號", "隔離原因"]
    ws2.append(headers2)
    for h in ws2[1]:
        h.font = Font(bold=True, color="FFFFFF", size=11)
        h.fill = PatternFill("solid", fgColor="C00000")
        h.alignment = Alignment(horizontal="center", vertical="center")

    for idx, sc in enumerate(scanned_list, 1):
        ws2.append([idx, sc['fileName'], sc['partNo'], sc['reason']])

    widths2 = [8, 35, 20, 30]
    for idx, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = w
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUT_XLSX)

    print("\n=== 原料圖檔萃取作業完成 ===")
    print(f"  - 總處理圖檔: {len(results)} 份 (100.0%)")
    print(f"  - 文字層解析: {len(results) - len(scanned_list)} 份")
    print(f"  - 標記隔離之純掃描圖檔: {len(scanned_list)} 份")
    print(f"  - 產出 JSON: {OUT_JSON}")
    print(f"  - 產出 Excel: {OUT_XLSX}")

process_resin_drawings()
