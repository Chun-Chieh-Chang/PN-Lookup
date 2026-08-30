import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open('data/assembly_deduped.json') as f:
    candidates = json.load(f)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '組件搬運清單'

# Styles
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='2E5090')
high_fill = PatternFill('solid', fgColor='FFEB9C')   # yellow
med_fill = PatternFill('solid', fgColor='DDEBF7')    # light blue
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = ['優先級', '序號', '檔名', '零件編號', '描述', '子零件清單', '目前路徑', '建議搬運至']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

ws.row_dimensions[1].height = 25

for i, c in enumerate(candidates, 1):
    fp = c['filePath'].replace('\\', '/')
    suggested = fp.replace('/零件/', '/組件/Component/')
    priority = 'HIGH' if c['score'] >= 9 else 'MED'
    fill = high_fill if c['score'] >= 9 else med_fill

    row = [
        priority,
        i,
        c['fileName'],
        c['partNo'],
        c['desc'],
        ', '.join(c['other_parts']),
        fp,
        suggested,
    ]
    for col, val in enumerate(row, 1):
        cell = ws.cell(row=i+1, column=col, value=val)
        cell.fill = fill
        cell.border = thin_border
        if col in (6, 7, 8):
            cell.alignment = Alignment(wrap_text=True)

# Column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 6
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 50
ws.column_dimensions['F'].width = 38
ws.column_dimensions['G'].width = 70
ws.column_dimensions['H'].width = 70

out = 'data/組件搬運清單.xlsx'
wb.save(out)
print(f'Saved {len(candidates)} candidates to {out}')
print(f'  HIGH (score>=9): {sum(1 for c in candidates if c["score"]>=9)}')
print(f'  MED  (score<9):  {sum(1 for c in candidates if c["score"]<9)}')
