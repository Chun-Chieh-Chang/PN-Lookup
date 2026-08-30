#!/usr/bin/env python3
"""Generate Excel from v7 JSON."""
import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

with open("data/drawings_extract_v7.json", encoding="utf-8") as f:
    d = json.load(f)
items = d["items"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "圖面資料"
headers = ["圖檔名", "圖號", "版本", "品號", "品名", "顏色", "原料名稱", "原料編碼", "分類", "提取方法", "來源資料夾", "檔案路徑"]
ws.append(headers)
for h in ws[1]:
    h.font = Font(bold=True, color="FFFFFF")
    h.fill = PatternFill("solid", fgColor="4472C4")
    h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for r in items:
    ws.append([
        r["fileName"], r.get("drawingNo",""), r.get("revision",""), r.get("partNo",""),
        (r.get("description","") or "")[:100], r.get("color","") or "",
        (r.get("materialName","") or "")[:120], r.get("materialCode","") or "",
        r.get("category","零件"), r.get("method",""), r.get("source",""), r.get("filePath","")
    ])

col_widths = [40, 22, 8, 20, 50, 12, 50, 20, 12, 14, 25, 60]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

scanned_set = set(r["fileName"] for r in items if r.get("method") == "scanned_no_text")
red_fill = PatternFill("solid", fgColor="FFE0E0")
for row in ws.iter_rows(min_row=2):
    if row[0].value in scanned_set:
        for cell in row:
            cell.fill = red_fill

ws2 = wb.create_sheet("組件BOM")
ws2.append(["組件圖檔名", "組件品號", "分類", "組成零件品號", "品名", "原料", "用量"])
for h in ws2[1]:
    h.font = Font(bold=True, color="FFFFFF")
    h.fill = PatternFill("solid", fgColor="4472C4")
for r in items:
    if r.get("bom"):
        for b in r["bom"]:
            ws2.append([r["fileName"], r.get("partNo","") or r.get("drawingNo",""), r.get("category","零件"),
                       b.get("partNo",""), b.get("description",""), b.get("material",""), b.get("qty","")])
for i, w in enumerate([35, 20, 10, 20, 40, 50, 8], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws3 = wb.create_sheet("掃描圖檔需手動處理")
ws3.append(["圖檔名", "狀態", "來源資料夾"])
for h in ws3[1]:
    h.font = Font(bold=True, color="FFFFFF")
    h.fill = PatternFill("solid", fgColor="C00000")
for r in items:
    if r.get("method") in ("scanned_no_text",):
        ws3.append([r["fileName"], "(純圖片無文字層)", r.get("source","")])
for i, w in enumerate([40, 25, 20], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

wb.save("data/drawings_extract_v7.xlsx")
print("Excel saved")
