import fitz  # PyMuPDF
import os, sys, io, re, json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT_DIR = r"D:\Self-developed_Apps\PN-Lookup\rawdata\Drawings\物料"
OUT_JSON = r"D:\Self-developed_Apps\PN-Lookup\data\material_drawings_extract.json"
OUT_XLSX = r"D:\Self-developed_Apps\PN-Lookup\data\material_drawings_extract.xlsx"

print(f"=== 啟動「物料」圖檔資訊萃取管線 ===")
print(f"掃描根目錄: {ROOT_DIR}")

# 顏色對應規則
COLOR_PATTERNS = [
    (r'土黃色|牛皮色', 'Kraft (土黃色/牛皮色)'),
    (r'透明藍|BLUE\s*TINT', 'Blue Tint (透明藍)'),
    (r'透明|CLEAR|TRANSPARENT', 'Clear / Transparent (透明)'),
    (r'白色|WHITE|白底', 'White (白)'),
    (r'黑色|BLACK|黑字', 'Black (黑)'),
    (r'紅色|RED', 'Red (紅)'),
    (r'藍色|BLUE', 'Blue (藍)'),
    (r'綠色|GREEN', 'Green (綠)'),
    (r'黃色|YELLOW', 'Yellow (黃)'),
    (r'四色印刷|彩色印刷|印刷', 'Printed (印刷彩色)'),
]

# 預設各子資料夾的常見材質（若圖面未明記）
DEFAULT_MATERIALS = {
    '外箱標籤 CL': '銅版貼紙 / 標籤紙',
    '內箱標籤 ICL': '銅版貼紙 / 標籤紙',
    '標籤 MXXXX': '標籤貼紙 (Label Paper)',
    '標籤貼紙 BL': '標籤貼紙 (Label Sticker)',
    '紙箱 CTN': '瓦楞紙板 (Corrugated Cardboard)',
    '包裝袋 PL': '透析紙/滅菌袋 (Tyvek / Film Pouch)',
    '塑膠袋 PE': 'PE 塑膠袋 (Polyethylene)',
    '收縮膜 SF': 'PVC/PET 收縮膜 (Shrink Film)',
    '包裝說明書 PN': '模造紙 / 道林紙 (Instruction Paper)',
    '說明書 PI': '模造紙 / 道林紙 (Manual Paper)',
}

def extract_from_filename(filename, folder_name):
    """從檔名解析備援資訊"""
    base = os.path.splitext(filename)[0]
    
    # 1. 提取版本
    rev = ""
    m_rev = re.search(r'(?:Rev|Rec|Ver|版次)[\._\s]*([A-Z0-9]+)', base, re.I)
    if m_rev:
        rev = f"Rev.{m_rev.group(1).upper()}"
    elif re.search(r'_A0[1-9]', base, re.I):
        m_a = re.search(r'_(A0[1-9])', base, re.I)
        rev = m_a.group(1)

    # 2. 提取品號與圖號
    part_no = ""
    dwg_no = ""
    # 常見物料編號前綴: CL, ICL, PL, CTN, PN, PE, SF, BL, PI, MXXXX (如 M0022)
    m_pn = re.search(r'(\bM\d{4}\b|[A-Z]{2,4}-?\d{3,5}(?:-\d+)?)', base, re.I)
    if m_pn:
        part_no = m_pn.group(1).upper()
        dwg_no = part_no
    elif re.search(r'9X\.\d{5}\.\d{3}', base):
        m_cust = re.search(r'9X\.\d{5}\.\d{3}', base)
        part_no = m_cust.group(0)
        m_inner = re.search(r'PL-\d{4}', base)
        dwg_no = m_inner.group(0) if m_inner else part_no

    # 3. 提取品名
    desc = ""
    if '_' in base:
        parts = base.split('_')
        desc = parts[-1].strip()
        # 清理結尾簽核字串
        desc = re.sub(r'-(?:signed|A0[1-9]).*', '', desc, flags=re.I).strip()
    elif '(' in base and ')' in base:
        # 取括號外或括號內中文
        m_desc = re.search(r'[\u4e00-\u9fa5]+[^\.]*', base)
        if m_desc: desc = m_desc.group(0).strip()

    if not desc:
        desc = base

    return part_no, dwg_no, rev, desc

def process_material_drawings():
    results = []
    scanned_list = []
    all_bom_rows = []

    for root, dirs, files in os.walk(ROOT_DIR):
        rel_dir = os.path.relpath(root, ROOT_DIR)
        folder_top = rel_dir.split(os.sep)[0] if rel_dir != '.' else '物料'
        
        for f in files:
            if not f.lower().endswith('.pdf'):
                continue
            
            fp = os.path.join(root, f)
            # 從檔名解析基礎屬性
            fn_pn, fn_dwg, fn_rev, fn_desc = extract_from_filename(f, folder_top)
            
            lines = []
            try:
                doc = fitz.open(fp)
                for page in doc:
                    lines.extend([l.strip() for l in page.get_text().splitlines() if l.strip()])
            except Exception as e:
                lines = []

            # 判斷是否為純掃描件
            is_scanned = (len(lines) <= 2)
            if is_scanned:
                scanned_list.append({
                    'fileName': f,
                    'filePath': fp,
                    'folder': rel_dir,
                    'partNo': fn_pn,
                    'reason': f"文字層行數不足 ({len(lines)} 行)"
                })

            full_text = " ".join(lines)

            # 1. 圖號與品號精化
            dwg_no = fn_dwg
            part_no = fn_pn
            m_pn_body = re.search(r'(?:PART\s*NO|零件編號|品號|圖號)[\.:\s]+([A-Z0-9\-_]{3,20})', full_text, re.I)
            if m_pn_body:
                cand = m_pn_body.group(1).upper()
                if not re.search(r'FILE|NONE|TOLERANCE|CKD|DESCRIPTION|REVISION|SCALE|UNIT|DATE', cand):
                    if not part_no:
                        part_no = cand
                        if not dwg_no: dwg_no = part_no

            # 2. 版本精化
            rev = fn_rev
            m_rev_body = re.search(r'(?:REV\.|REVISION|版次)[\.:\s]+([A-Z0-9]+)', full_text, re.I)
            if m_rev_body:
                r_cand = m_rev_body.group(1).upper()
                if len(r_cand) <= 4 and not re.search(r'NONE|PAGE|DATE', r_cand):
                    rev = f"Rev.{r_cand}"

            # 3. 品名精化
            desc = fn_desc
            m_desc_body = re.search(r'(?:DESCRIPTION|零件名稱|品名)[\.:\s]+([^\n\r\|]{3,40})', full_text, re.I)
            if m_desc_body:
                d_cand = m_desc_body.group(1).strip()
                if not re.search(r'FILE|PART|CKD|REV|DATE', d_cand):
                    desc = d_cand

            # 4. 顏色
            color = ""
            for pat, cname in COLOR_PATTERNS:
                if re.search(pat, full_text, re.I) or re.search(pat, desc, re.I):
                    color = cname
                    break

            # 5. 原料材質 (Material)
            material = ""
            m_mat = re.search(r'(?:MATERIAL|材\s*質)[\.:\s]+([^\n\r\|]{2,40})', full_text, re.I)
            if m_mat:
                m_cand = m_mat.group(1).strip()
                if not re.search(r'TOLERANCES|FINISHED|SCALE|UNIT|REVISION|NONE|N/A', m_cand, re.I):
                    material = m_cand

            if not material:
                # 檢查是否有注意事項中的材質標註
                m_spec = re.search(r'規格[\.:\s]+([^\n\r\|]{2,30})', full_text)
                if m_spec:
                    material = m_spec.group(1).strip()
                else:
                    # 使用子目錄之領域知識預設
                    material = DEFAULT_MATERIALS.get(folder_top, '包裝物料')

            # 6. 原料編碼 (Material Code)
            mat_code = ""
            m_code = re.search(r'(?:原料編碼|料號|材料編號)[\.:\s]+([A-Z0-9\-_]{4,20})', full_text, re.I)
            if m_code:
                mat_code = m_code.group(1).strip()

            # 7. 物料類別 (Category)
            # 依據規則定義：包含"標籤"、"包裝袋"、"說明書"、"收縮膜"、"紙箱"、"塑膠袋"、"標籤貼紙"等皆屬之
            category = "物料"

            # 8. BOM 明細抽取 (若為包裝組合或含多物料)
            bom_details = []
            if part_no and len(part_no) >= 3:
                m_boms = re.findall(r'\b(CTN\d{3}|ICL-\d{4}|CL-\d{4}|PL-\d{4}|SF\d{3}|PE\d{3}|0\.08[xX*]\d+mm)\b', full_text)
                seen_boms = set()
                for b_pn in m_boms:
                    b_pn = b_pn.upper()
                    if b_pn != part_no and b_pn not in seen_boms:
                        seen_boms.add(b_pn)
                        bom_details.append({
                            'qty': '1',
                            'partNo': b_pn,
                            'name': b_pn,
                            'material': '包裝物料',
                            'materialCode': ''
                        })
                        all_bom_rows.append({
                            'fileName': f,
                            'parentPartNo': part_no,
                            'qty': '1',
                            'childPartNo': b_pn,
                            'childName': b_pn,
                            'childMaterial': '包裝物料'
                        })

            results.append({
                'fileName': f,
                'dwgNo': dwg_no,
                'revision': rev,
                'partNo': part_no,
                'name': desc,
                'color': color,
                'material': material,
                'materialCode': mat_code,
                'category': category,
                'folder': rel_dir,
                'isScanned': is_scanned,
                'bomDetails': bom_details
            })

    # 輸出 JSON
    out_payload = {
        'totalDrawings': len(results),
        'textLayerCount': len(results) - len(scanned_list),
        'scannedCount': len(scanned_list),
        'bomRowsCount': len(all_bom_rows),
        'parts': results,
        'scannedList': scanned_list,
        'bomRows': all_bom_rows
    }
    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(out_payload, f, ensure_ascii=False, indent=2)

    # 輸出 Excel
    wb = openpyxl.Workbook()
    
    # 工作表 1: 物料圖檔規格總表
    ws1 = wb.active
    ws1.title = "物料圖檔規格總表"
    headers1 = [
        "1.圖檔檔名", "2.圖號(Drawing No.)", "3.版本(REV.)", "4.品號(Part No.)",
        "5.品名(Description)", "6.顏色(Color)", "7.原料名稱(Material)",
        "8.原料編碼(Material Code)", "9.物料類別(Category)", "子目錄資料夾", "解析方式"
    ]
    ws1.append(headers1)
    for h in ws1[1]:
        h.font = Font(bold=True, color="FFFFFF", size=11)
        h.fill = PatternFill("solid", fgColor="1F497D")
        h.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        method = "待OCR處理 (純掃描件)" if r['isScanned'] else "文字層精準解析"
        ws1.append([
            r['fileName'], r['dwgNo'], r['revision'], r['partNo'],
            r['name'], r['color'], r['material'], r['materialCode'],
            r['category'], r['folder'], method
        ])

    widths1 = [38, 20, 12, 20, 38, 18, 30, 18, 14, 25, 20]
    for idx, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(idx)].width = w
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 工作表 2: 物料BOM組成清單
    ws2 = wb.create_sheet("物料BOM組成清單")
    headers2 = ["圖檔檔名", "母件品號", "單位用量", "組成物料品號", "組成物料品名", "物料材質"]
    ws2.append(headers2)
    for h in ws2[1]:
        h.font = Font(bold=True, color="FFFFFF", size=11)
        h.fill = PatternFill("solid", fgColor="203764")
        h.alignment = Alignment(horizontal="center", vertical="center")

    for b in all_bom_rows:
        ws2.append([b['fileName'], b['parentPartNo'], b['qty'], b['childPartNo'], b['childName'], b['childMaterial']])

    widths2 = [35, 18, 12, 20, 35, 25]
    for idx, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = w
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 工作表 3: 待OCR掃描圖檔清單
    ws3 = wb.create_sheet("待OCR掃描圖檔清單")
    headers3 = ["項次", "圖檔檔名", "所在資料夾", "檔名推導品號", "隔離原因"]
    ws3.append(headers3)
    for h in ws3[1]:
        h.font = Font(bold=True, color="FFFFFF", size=11)
        h.fill = PatternFill("solid", fgColor="C00000")
        h.alignment = Alignment(horizontal="center", vertical="center")

    for idx, sc in enumerate(scanned_list, 1):
        ws3.append([idx, sc['fileName'], sc['folder'], sc['partNo'], sc['reason']])

    widths3 = [8, 38, 25, 20, 30]
    for idx, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(idx)].width = w
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUT_XLSX)

    print("\n=== 物料圖檔萃取作業完成 ===")
    print(f"  - 總處理圖檔: {len(results)} 份 (100.0%)")
    print(f"  - 文字層解析圖檔: {len(results) - len(scanned_list)} 份")
    print(f"  - 標記隔離之純掃描圖檔: {len(scanned_list)} 份")
    print(f"  - 展開物料組成關聯: {len(all_bom_rows)} 行")
    print(f"  - 產出 JSON: {OUT_JSON}")
    print(f"  - 產出 Excel: {OUT_XLSX}")

process_material_drawings()
