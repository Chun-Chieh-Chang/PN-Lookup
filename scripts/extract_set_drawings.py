import os, sys, fitz, json, re, io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

print("=== 啟動 SET 圖面 PDF 結構化數據全量提取作業 (SET Drawing Pipeline) ===")

BASE_DIR = r"D:\Self-developed_Apps\PN-Lookup\rawdata\Drawings\SET"
JSON_OUTPUT = r"D:\Self-developed_Apps\PN-Lookup\data\set_drawings_extract.json"
XLSX_OUTPUT = r"D:\Self-developed_Apps\PN-Lookup\data\set_drawings_extract.xlsx"

# 1. 載入主資料庫作為交叉比對與補強基準
with open('data/pn-lookup-master.json', 'r', encoding='utf-8') as f:
    master = json.load(f)

master_parts = {}
for p in master.get('parts', []):
    pn = p['partNo'].strip().upper()
    master_parts[pn] = p
    for a in p.get('alternates', []):
        master_parts[a.strip().upper()] = p

COLOR_KEYWORDS = [
    (r'\bBLUE\s+TINT\b|透明藍', 'Blue Tint (透明藍)'),
    (r'\bOPAQUE\s+BLUE\b|不透明藍', 'Opaque Blue (不透明藍)'),
    (r'\bLIGHT\s+BLUE\b|淺藍', 'Light Blue (淺藍)'),
    (r'\bDARK\s+BLUE\b|深藍', 'Dark Blue (深藍)'),
    (r'\bOPAQUE\s+WHITE\b|不透明白', 'Opaque White (不透明白)'),
    (r'\bOFF-WHITE\b|米白', 'Off-White (米白)'),
    (r'\bDARK\s+GREEN\b|深綠', 'Dark Green (深綠)'),
    (r'\bLIGHT\s+GREEN\b|淺綠', 'Light Green (淺綠)'),
    (r'\bWHITE\b|\bWHITE\s+PC\b|白', 'White (白)'),
    (r'\bBLACK\b|黑', 'Black (黑)'),
    (r'\bBLUE\b|\bBLUE\s+PC\b|藍', 'Blue (藍)'),
    (r'\bRED\b|紅', 'Red (紅)'),
    (r'\bGREEN\b|綠', 'Green (綠)'),
    (r'\bCLEAR\b|透明|\bTRANS\b|\bTRANS[-_]?\d+', 'Clear / Transparent (透明)'),
    (r'\bNATURAL\b|本色|原色', 'Natural (本色/原色)'),
    (r'\bTRANSPARENT\b', 'Transparent (透明)'),
    (r'\bYELLOW\b|黃', 'Yellow (黃)'),
    (r'\bW[-_]?\d{3}\b', 'White (白色母)'),
    (r'\bK[-_]?\d{3}\b', 'Blue (藍色母)'),
]

PN_TOKEN_RE = re.compile(r'\b(?:[A-Z]{1,4}\d{1,4}(?:-\d{1,4}){1,3}[A-Z0-9]?|[A-Z]{2,4}\d{4,7}|\d{1,2}[A-Z]\d{3,6}|\d{4,}(?:-\d+)*|\d{2,3}(?:-\d+){1,3}|B-\d{3}|0\.08[xX*]\d+(?:\.\d+)?mm?)\b', re.I)

def extract_set_metadata(filename, text=""):
    stem = os.path.splitext(filename)[0]
    
    # 1. MDXE / MDXI (如 MDXE-004-02_A -> partNo: MDXE-004-02, rev: A, dwg: MDXE-004-02)
    m = re.match(r'^(MDX[EI]-[\w\-]+)_([A-Z0-9]+)$', stem, re.I)
    if m:
        return m.group(1).upper(), m.group(2).upper(), m.group(1).upper()
        
    # 2. MPS(LSO) 格式 (如 EB03002(Rev.B) -> partNo: EB03002, rev: B, dwg: EB03002)
    m = re.search(r'^([A-Z]{2}\d{5}[A-Z0-9]*)\((?:Rev\.?)?([A-Z0-9]+)\)', stem, re.I)
    if m:
        return m.group(1).upper(), m.group(2).upper(), m.group(1).upper()
        
    # 3. BD 格式 (如 BD-8003875_Rev.04 -> partNo: 8003875, rev: 04, dwg: BD-8003875)
    m = re.search(r'BD[-_]?([A-Z0-9]+)_Rev\.?([A-Z0-9]+)', stem, re.I)
    if m:
        return m.group(1).upper(), m.group(2).upper(), f"BD-{m.group(1).upper()}"

    # 4. 其他包含 (Rev.X) 格式
    m_rev = re.search(r'\((?:Rev\.?)?([A-Z0-9]+)\)', stem, re.I)
    rev = m_rev.group(1).upper() if m_rev else ""
    
    # 提取品號
    m_pn = re.search(r'([A-Z0-9\-]+)', stem)
    pn = m_pn.group(1).upper() if m_pn else stem
    return pn, rev, pn

def extract_set_text_bom(text, part_no):
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    bom_items = []
    
    # 模式 A: MDXE 表格 (Item -> Q'ty -> P/N -> Description -> Material code -> Material)
    # 尋找 Item 表頭
    item_header_idx = -1
    for i, l in enumerate(lines):
        if re.search(r'Item\b|P/N\b|Q[\'’]?ty\b', l, re.I):
            item_header_idx = i
            break
            
    if item_header_idx != -1:
        # 尋找由純數字構成的 Item 序列 (1, 2, 3...)
        # 在 MDXE-019-03 中，Item 序列先出現 (1, 2, 3, 4, 5)，接著是 Qty (1, 1, 1...)，接著是 P/N
        # 或是每行並列
        for i in range(item_header_idx, min(len(lines), item_header_idx + 80)):
            line = lines[i]
            # 檢查是否為品號
            if PN_TOKEN_RE.search(line) and not re.search(r'ISO|PER|REV|DATE|SCALE|TOLERANCE', line, re.I):
                pn_token = PN_TOKEN_RE.search(line).group(0)
                if pn_token.upper() != part_no.upper() and len(pn_token) >= 4:
                    # 尋找前後上下文
                    desc = ""
                    mat = ""
                    # 往下看 1~3 行
                    for nxt in lines[i+1 : min(len(lines), i+4)]:
                        if not desc and len(nxt) >= 3 and not PN_TOKEN_RE.search(nxt):
                            desc = nxt
                        elif not mat and re.search(r'PVC|HDPE|PP|ABS|PC|SILICONE|RUBBER|7088G', nxt, re.I):
                            mat = nxt
                    bom_items.append({
                        'qty': '1',
                        'partNo': pn_token,
                        'name': desc or pn_token,
                        'material': mat,
                        'materialCode': ''
                    })

    # 模式 B: BD 輸液套表格 (ITEM | QTY | PART NO | DESCRIPTION)
    start_bd = -1
    for i, l in enumerate(lines):
        if re.search(r'BILL\s+OF\s+MATERIALS|ITEM\s+QTY\s+PART\s+NO', l, re.I):
            start_bd = i
            break
            
    if start_bd != -1:
        for i in range(start_bd, min(len(lines), start_bd + 60)):
            line = lines[i]
            m_pn = PN_TOKEN_RE.search(line)
            if m_pn:
                pn_token = m_pn.group(0)
                if pn_token.upper() != part_no.upper() and len(pn_token) >= 4:
                    desc = line.replace(pn_token, '').strip()
                    if not desc and i + 1 < len(lines):
                        desc = lines[i+1]
                    bom_items.append({
                        'qty': '1',
                        'partNo': pn_token,
                        'name': desc or pn_token,
                        'material': 'PVC' if 'TUBING' in (desc or '').upper() else '',
                        'materialCode': ''
                    })

    # 去重
    seen = set()
    clean = []
    for b in bom_items:
        cpn = b['partNo'].upper()
        if cpn not in seen and cpn != part_no.upper():
            seen.add(cpn)
            clean.append(b)
            
    return clean

# ─────────────────────────────────────────────────────────────
# 遍歷 113 個 SET PDF 檔案
# ─────────────────────────────────────────────────────────────
all_items = []
scanned_items = []
all_bom_rows = []

pdf_list = []
for root, dirs, files in os.walk(BASE_DIR):
    for f in files:
        if f.lower().endswith(".pdf"):
            pdf_list.append(os.path.join(root, f))

pdf_list.sort()
print(f"開始解析 {len(pdf_list)} 份 SET PDF 圖檔...")

for p in pdf_list:
    fname = os.path.basename(p)
    rel_folder = os.path.relpath(os.path.dirname(p), BASE_DIR)
    
    text = ""
    is_scanned = False
    try:
        doc = fitz.open(p)
        for page in doc:
            text += page.get_text() + "\n"
        doc.close()
        if len(text.strip()) < 40:
            is_scanned = True
    except Exception as e:
        is_scanned = True

    # 1. 提取核心屬性
    part_no, revision, drawing_no = extract_set_metadata(fname, text)
    
    # 從 master 取得對照實體
    target_pn = 'X3299AAM' if part_no == 'X3299' else part_no
    m_part = master_parts.get(target_pn.upper(), {})
    
    # 品名 (Description)
    description = ""
    if m_part.get('name') and m_part['name'] != m_part['partNo']:
        description = m_part['name']
    elif m_part.get('description'):
        description = m_part['description']
    else:
        # 從文字層抓取
        m_desc = re.search(r'(?:DESCRIPTION|TITLE)[\.:\s]+([^\n\r\|]{3,60})', text, re.I)
        if m_desc:
            description = m_desc.group(1).strip()
        elif 'MDX' in part_no:
            description = '輸液套延長管 (Extension Set)'
            
    # 原料名稱 (Material)
    material_name = m_part.get('material') or ''
    if not material_name or material_name in ('零件', '組件', 'N/A', 'NONE'):
        m_mat = re.search(r'(?:MATERIAL|材\s*質)[\.:\s]+([^\n\r\|]{3,60})', text, re.I)
        if m_mat:
            material_name = m_mat.group(1).strip()
        elif 'MDX' in part_no:
            material_name = 'PVC, DEHP FREE'
            
    # 原料編碼 (Material Code)
    material_code = m_part.get('materialCode') or ''
    
    # 2. 提取子零件 BOM 清單
    bom_details = []
    # (A) 文字層提取
    if not is_scanned:
        bom_details = extract_set_text_bom(text, part_no)
        
    # (B) Master 回退保底
    if not bom_details and m_part.get('bomDetails'):
        bom_details = list(m_part['bomDetails'])
    elif not bom_details and target_pn.upper() in master.get('bom', {}).get('children', {}):
        for cpn in master['bom']['children'][target_pn.upper()]:
            cp = master_parts.get(cpn.upper(), {})
            bom_details.append({
                'qty': '1',
                'partNo': cpn,
                'name': cp.get('name') or cp.get('description') or cpn,
                'material': cp.get('material') or '',
                'materialCode': cp.get('materialCode') or ''
            })

    # 去重
    seen_bom = set()
    clean_boms = []
    for b in bom_details:
        cpn = b['partNo'].strip().upper()
        if cpn not in seen_bom and cpn != part_no.upper():
            seen_bom.add(cpn)
            clean_boms.append(b)
            all_bom_rows.append({
                'setFile': fname,
                'setPartNo': part_no,
                'setName': description,
                'qty': b.get('qty', '1'),
                'childPartNo': b['partNo'],
                'childName': b.get('name', ''),
                'childMaterial': b.get('material', ''),
                'childMaterialCode': b.get('materialCode', '')
            })

    # 3. 顏色提取
    color = m_part.get('color') or ''
    if not color:
        bom_str = " ".join((b.get('name', '') + " " + b.get('material', '')) for b in clean_boms)
        comb = f"{text} {material_name} {description} {bom_str}"
        for pat, cname in COLOR_KEYWORDS:
            if re.search(pat, comb, re.I):
                color = cname
                break
                
    # 4. 物料類別: 100% 歸為 SET
    category = "SET"
    
    item_record = {
        'fileName': fname,
        'drawingNo': drawing_no,
        'revision': revision,
        'partNo': part_no,
        'description': description,
        'color': color,
        'materialName': material_name,
        'materialCode': material_code,
        'category': category,
        'bomDetails': clean_boms,
        'isScanned': is_scanned,
        'folder': rel_folder,
        'filePath': p
    }
    all_items.append(item_record)
    
    if is_scanned:
        scanned_items.append({
            'fileName': fname,
            'partNo': part_no,
            'folder': rel_folder,
            'filePath': p,
            'status': '純掃描圖檔 (文字層無內容，留待OCR處理)'
        })

print(f"\nSET 圖面提取統計成果:")
print(f"  - 總圖檔筆數: {len(all_items)} 筆")
print(f"  - 品號填寫率: {sum(1 for i in all_items if i['partNo'])} / {len(all_items)} (100.0%)")
print(f"  - 圖號填寫率: {sum(1 for i in all_items if i['drawingNo'])} / {len(all_items)} (100.0%)")
print(f"  - 版本填寫率: {sum(1 for i in all_items if i['revision'])} / {len(all_items)} ({sum(1 for i in all_items if i['revision'])/len(all_items)*100:5.1f}%)")
print(f"  - 顏色填寫率: {sum(1 for i in all_items if i['color'])} / {len(all_items)} ({sum(1 for i in all_items if i['color'])/len(all_items)*100:5.1f}%)")
print(f"  - 組成零件記錄筆數 (BOM 展開行): 共 {len(all_bom_rows)} 行明細")
print(f"  - 掃描圖檔標記筆數: 共 {len(scanned_items)} 筆")

# 儲存 JSON
with open(JSON_OUTPUT, 'w', encoding='utf-8') as f:
    json.dump({
        'totalDrawings': len(all_items),
        'totalBomRows': len(all_bom_rows),
        'scannedCount': len(scanned_items),
        'items': all_items,
        'scannedList': scanned_items
    }, f, ensure_ascii=False, indent=2)
print(f"已儲存 JSON 成果: {JSON_OUTPUT}")

# ─────────────────────────────────────────────────────────────
# 產出標準 Excel 文件: set_drawings_extract.xlsx
# ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# 工作表 1: SET圖面資料 (9 大核心欄位)
ws1 = wb.active
ws1.title = "SET圖面資料"
headers1 = [
    "1.圖檔檔名", "2.圖號(Drawing No.)", "3.版本(REV.)", "4.品號(Part No.)",
    "5.品名(Description)", "6.顏色(Color)", "7.原料名稱(Material)",
    "8.原料編碼(Material Code)", "9.物料類別(Category)", "子件數量", "圖檔類型", "資料夾目錄"
]
ws1.append(headers1)
for h in ws1[1]:
    h.font = Font(bold=True, color="FFFFFF", size=11)
    h.fill = PatternFill("solid", fgColor="1F497D")
    h.alignment = Alignment(horizontal="center", vertical="center")

for it in all_items:
    ws1.append([
        it['fileName'], it['drawingNo'], it['revision'], it['partNo'],
        it['description'], it['color'], it['materialName'], it['materialCode'],
        it['category'], len(it['bomDetails']),
        "掃描圖檔(待OCR)" if it['isScanned'] else "文字層向量PDF",
        it['folder']
    ])

widths1 = [38, 22, 10, 18, 42, 16, 40, 18, 14, 10, 18, 25]
for idx, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(idx)].width = w
for row in ws1.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# 工作表 2: SET_BOM清單
ws2 = wb.create_sheet("SET_BOM清單")
headers2 = [
    "SET圖檔名", "SET品號", "SET品名", "單位用量(Qty)",
    "組成零件品號", "組成零件品名", "原料名稱(Material)", "原料編碼(Material Code)"
]
ws2.append(headers2)
for h in ws2[1]:
    h.font = Font(bold=True, color="FFFFFF", size=11)
    h.fill = PatternFill("solid", fgColor="203764")
    h.alignment = Alignment(horizontal="center", vertical="center")

for b in all_bom_rows:
    ws2.append([
        b['setFile'], b['setPartNo'], b['setName'],
        b['qty'], b['childPartNo'], b['childName'],
        b['childMaterial'], b['childMaterialCode']
    ])

widths2 = [35, 18, 35, 14, 18, 38, 38, 18]
for idx, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(idx)].width = w
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# 工作表 3: 掃描圖檔清單(待OCR)
ws3 = wb.create_sheet("掃描圖檔清單(待OCR)")
headers3 = ["圖檔檔名", "品號", "來源資料夾", "檔案完整路徑", "處理狀態備註"]
ws3.append(headers3)
for h in ws3[1]:
    h.font = Font(bold=True, color="FFFFFF", size=11)
    h.fill = PatternFill("solid", fgColor="C00000")
    h.alignment = Alignment(horizontal="center", vertical="center")

for sf in scanned_items:
    ws3.append([
        sf['fileName'], sf['partNo'], sf['folder'], sf['filePath'], sf['status']
    ])

widths3 = [38, 18, 25, 65, 35]
for idx, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(idx)].width = w
for row in ws3.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(XLSX_OUTPUT)
print(f"已成功重新生成高品質 Excel 文件: {XLSX_OUTPUT}")
